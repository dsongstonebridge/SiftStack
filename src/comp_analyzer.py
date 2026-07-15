"""Comparable sales analysis with Two-Bucket ARV methodology.

Generates appraiser-grade property valuations for real estate investment
analysis. Fetches comparable sales from the Zillow API, applies property-
specific adjustments, and produces a 7-tab Excel workbook.

Tennessee is a non-disclosure state — MLS/Zillow data is the primary
source, not public deed records.

Usage:
  python src/main.py comp --address "123 Main St, Knoxville, TN 37918"
  python src/main.py comp --address "123 Main St" --city Knoxville --zip 37918 --radius 0.5 --months 6
"""

import logging
import math
import random
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config

logger = logging.getLogger(__name__)

# ── API Configuration ─────────────────────────────────────────────────
API_BASE = "https://api.openwebninja.com/realtime-zillow-data"
PROPERTY_ENDPOINT = f"{API_BASE}/property-details-address"
COMPS_ENDPOINT = f"{API_BASE}/similar-sale-homes"
REQUEST_DELAY_MIN = 1.0
REQUEST_DELAY_MAX = 2.0
REQUEST_TIMEOUT = 30
MAX_RETRIES = 2

# ── Comp selection defaults ───────────────────────────────────────────
DEFAULT_RADIUS_MILES = 0.5
MAX_RADIUS_MILES = 1.0
DEFAULT_MONTHS_BACK = 6
MAX_MONTHS_BACK = 12
MIN_COMPS = 3
TARGET_COMPS = 5
MAX_COMPS = 7

# ── Adjustment values (Knoxville regional calibration) ────────────────
# These are per-unit adjustment amounts used when a comp differs from subject
ADJ_PER_SQFT = 85.0            # $ per sqft difference
ADJ_PER_BEDROOM = 5000.0       # $ per bedroom difference
ADJ_PER_BATHROOM = 7500.0      # $ per bathroom difference
ADJ_PER_YEAR_BUILT = 500.0     # $ per year of age difference
ADJ_PER_LOT_SQFT = 2.0         # $ per sqft of lot size difference
ADJ_LOT_MAX = 15000.0          # Cap on lot size adjustment
ADJ_PER_GARAGE = 8000.0        # $ per garage stall difference
# Market condition adjustment: % per month of age (appreciating market)
MARKET_CONDITION_PCT_PER_MONTH = 0.003  # 0.3% per month ≈ 3.6% annual

# ── Data structures ───────────────────────────────────────────────────


@dataclass
class SubjectProperty:
    """The property being analyzed."""
    address: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    latitude: float = 0.0
    longitude: float = 0.0
    sqft: int = 0
    bedrooms: int = 0
    bathrooms: float = 0.0
    year_built: int = 0
    lot_sqft: int = 0
    property_type: str = ""
    zestimate: float = 0.0
    mls_status: str = ""
    last_sold_date: str = ""
    last_sold_price: float = 0.0
    garage_spaces: int = 0
    description: str = ""


@dataclass
class CompProperty:
    """A comparable property with sale data."""
    address: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    latitude: float = 0.0
    longitude: float = 0.0
    distance_miles: float = 0.0
    sqft: int = 0
    bedrooms: int = 0
    bathrooms: float = 0.0
    year_built: int = 0
    lot_sqft: int = 0
    property_type: str = ""
    sold_price: float = 0.0
    sold_date: str = ""
    days_on_market: int = 0
    garage_spaces: int = 0
    # Calculated fields
    similarity_score: float = 0.0
    adjusted_price: float = 0.0
    ppsf: float = 0.0
    bucket: str = ""  # "A" (non-disclosure baseline) or "B" (disclosure/adjusted)
    adjustments: dict = field(default_factory=dict)


@dataclass
class ARVResult:
    """Final ARV calculation result."""
    arv_low: float = 0.0
    arv_mid: float = 0.0
    arv_high: float = 0.0
    confidence: str = ""  # "high", "medium", "low"
    confidence_reason: str = ""
    ppsf_avg: float = 0.0
    ppsf_range: tuple = (0.0, 0.0)
    comp_count: int = 0
    bucket_a_count: int = 0
    bucket_b_count: int = 0
    avg_adjustment: float = 0.0
    spread_pct: float = 0.0


# ── Distance calculation ──────────────────────────────────────────────

def _haversine_miles(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate distance between two lat/lon points in miles."""
    R = 3958.8  # Earth radius in miles
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))


# ── API calls ─────────────────────────────────────────────────────────

def _api_get(endpoint: str, params: dict, api_key: str) -> dict | None:
    """Make an authenticated GET request to OpenWeb Ninja API."""
    headers = {"x-api-key": api_key}
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(endpoint, headers=headers, params=params, timeout=REQUEST_TIMEOUT)
            if resp.status_code == 404:
                return None
            if resp.status_code == 429:
                logger.warning("Rate limit hit — waiting 10s (attempt %d)", attempt)
                time.sleep(10)
                continue
            resp.raise_for_status()
            body = resp.json()
            if body.get("status") == "OK" and body.get("data"):
                return body["data"]
            return body if isinstance(body, list) else None
        except requests.Timeout:
            logger.warning("Timeout (attempt %d/%d)", attempt, MAX_RETRIES)
        except requests.RequestException as e:
            logger.warning("API error: %s (attempt %d/%d)", e, attempt, MAX_RETRIES)
    return None


def fetch_subject_property(address: str, city: str = "", state: str = "",
                           zip_code: str = "", api_key: str = "") -> SubjectProperty | None:
    """Fetch full property details for the subject property."""
    api_key = api_key or config.OPENWEBNINJA_API_KEY
    if not api_key:
        logger.error("No OpenWeb Ninja API key configured")
        return None

    parts = [p for p in [address, city, state, zip_code] if p]
    full_address = " ".join(parts)

    data = _api_get(PROPERTY_ENDPOINT, {"address": full_address}, api_key)
    if not data:
        logger.warning("No property data found for '%s'", full_address)
        return None

    # Parse price history for last sold
    last_sold_date, last_sold_price = "", 0.0
    for entry in (data.get("priceHistory") or []):
        event = (entry.get("event") or "").lower()
        if event in ("sold", "listed (sold)"):
            last_sold_date = str(entry.get("date", ""))[:10]
            last_sold_price = float(entry.get("price") or 0)
            break

    # Parse lot size
    lot_sqft = 0
    lot_val = data.get("lotAreaValue")
    lot_units = (data.get("lotAreaUnits") or data.get("lotAreaUnit") or "").lower()
    if lot_val:
        lot_sqft = int(float(lot_val) * 43560) if "acre" in lot_units else int(float(lot_val))

    return SubjectProperty(
        address=data.get("streetAddress") or address,
        city=data.get("city") or city,
        state=data.get("state") or state,
        zip_code=str(data.get("zipcode") or zip_code),
        latitude=float(data.get("latitude") or 0),
        longitude=float(data.get("longitude") or 0),
        sqft=int(data.get("livingArea") or 0),
        bedrooms=int(data.get("bedrooms") or 0),
        bathrooms=float(data.get("bathrooms") or 0),
        year_built=int(data.get("yearBuilt") or 0),
        lot_sqft=lot_sqft,
        property_type=data.get("homeType") or "",
        zestimate=float(data.get("zestimate") or 0),
        mls_status=data.get("homeStatus") or "",
        last_sold_date=last_sold_date,
        last_sold_price=last_sold_price,
        garage_spaces=int(data.get("garageSpaces") or 0),
        description=data.get("description") or "",
    )


def fetch_comparable_sales(subject: SubjectProperty, radius_miles: float = DEFAULT_RADIUS_MILES,
                           months_back: int = DEFAULT_MONTHS_BACK,
                           api_key: str = "") -> list[CompProperty]:
    """Fetch comparable sold properties near the subject property."""
    api_key = api_key or config.OPENWEBNINJA_API_KEY
    if not api_key:
        return []

    full_address = f"{subject.address} {subject.city} {subject.state} {subject.zip_code}"
    data = _api_get(COMPS_ENDPOINT, {"address": full_address}, api_key)

    time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))

    comps = []
    items = data if isinstance(data, list) else (data.get("comps") or data.get("results") or []) if data else []

    cutoff_date = (datetime.now() - timedelta(days=months_back * 30)).strftime("%Y-%m-%d")

    for item in items:
        if isinstance(item, str):
            continue

        # Parse sold info
        sold_price = float(item.get("lastSoldPrice") or item.get("price") or 0)
        sold_date = str(item.get("lastSoldDate") or item.get("dateSold") or "")[:10]

        if not sold_price or sold_price < 10000:
            continue

        # Filter by date
        if sold_date and sold_date < cutoff_date:
            continue

        lat = float(item.get("latitude") or 0)
        lon = float(item.get("longitude") or 0)

        # Filter by distance
        dist = 0.0
        if subject.latitude and subject.longitude and lat and lon:
            dist = _haversine_miles(subject.latitude, subject.longitude, lat, lon)
            if dist > radius_miles:
                continue

        lot_sqft = 0
        lot_val = item.get("lotAreaValue") or item.get("lotSize")
        lot_units = (item.get("lotAreaUnits") or "").lower()
        if lot_val:
            try:
                lot_sqft = int(float(lot_val) * 43560) if "acre" in lot_units else int(float(lot_val))
            except (ValueError, TypeError):
                pass

        comp = CompProperty(
            address=item.get("streetAddress") or item.get("address") or "",
            city=item.get("city") or "",
            state=item.get("state") or "",
            zip_code=str(item.get("zipcode") or item.get("zip") or ""),
            latitude=lat,
            longitude=lon,
            distance_miles=round(dist, 2),
            sqft=int(item.get("livingArea") or item.get("sqft") or 0),
            bedrooms=int(item.get("bedrooms") or 0),
            bathrooms=float(item.get("bathrooms") or 0),
            year_built=int(item.get("yearBuilt") or 0),
            lot_sqft=lot_sqft,
            property_type=item.get("homeType") or item.get("propertyType") or "",
            sold_price=sold_price,
            sold_date=sold_date,
            days_on_market=int(item.get("daysOnZillow") or 0),
            garage_spaces=int(item.get("garageSpaces") or 0),
        )
        comp.ppsf = round(comp.sold_price / comp.sqft, 2) if comp.sqft else 0.0
        comps.append(comp)

    logger.info("Fetched %d comparable sales within %.1f mi (last %d months)", len(comps), radius_miles, months_back)

    # If we don't have enough comps, try expanding radius and time window
    if len(comps) < MIN_COMPS and (radius_miles < MAX_RADIUS_MILES or months_back < MAX_MONTHS_BACK):
        new_radius = min(radius_miles * 1.5, MAX_RADIUS_MILES)
        new_months = min(months_back + 3, MAX_MONTHS_BACK)
        if new_radius > radius_miles or new_months > months_back:
            logger.info("Only %d comps — expanding search to %.1f mi / %d months",
                        len(comps), new_radius, new_months)
            # Re-filter with expanded params (comps already fetched, just relax filters)
            expanded_cutoff = (datetime.now() - timedelta(days=new_months * 30)).strftime("%Y-%m-%d")
            # Re-process original items with relaxed filters
            # For now, the API returns what it returns — we just note the expansion
            pass

    return comps


# ── Similarity scoring ────────────────────────────────────────────────

def _score_similarity(subject: SubjectProperty, comp: CompProperty) -> float:
    """Score how similar a comp is to the subject (0.0 to 1.0, higher = more similar)."""
    score = 1.0
    penalties = []

    # Square footage (most important)
    if subject.sqft and comp.sqft:
        sqft_diff_pct = abs(subject.sqft - comp.sqft) / subject.sqft
        if sqft_diff_pct > 0.30:
            score -= 0.30
            penalties.append(f"sqft {sqft_diff_pct:.0%} diff")
        elif sqft_diff_pct > 0.20:
            score -= 0.15
        elif sqft_diff_pct > 0.10:
            score -= 0.05

    # Bedrooms
    if subject.bedrooms and comp.bedrooms:
        bed_diff = abs(subject.bedrooms - comp.bedrooms)
        if bed_diff > 2:
            score -= 0.25
        elif bed_diff > 1:
            score -= 0.10
        elif bed_diff == 1:
            score -= 0.03

    # Bathrooms
    if subject.bathrooms and comp.bathrooms:
        bath_diff = abs(subject.bathrooms - comp.bathrooms)
        if bath_diff > 2:
            score -= 0.20
        elif bath_diff > 1:
            score -= 0.08

    # Year built
    if subject.year_built and comp.year_built:
        age_diff = abs(subject.year_built - comp.year_built)
        if age_diff > 20:
            score -= 0.20
        elif age_diff > 10:
            score -= 0.08
        elif age_diff > 5:
            score -= 0.03

    # Distance (closer = better)
    if comp.distance_miles > 0.75:
        score -= 0.15
    elif comp.distance_miles > 0.5:
        score -= 0.08
    elif comp.distance_miles > 0.25:
        score -= 0.03

    # Property type mismatch
    if subject.property_type and comp.property_type:
        if subject.property_type.upper() != comp.property_type.upper():
            score -= 0.20

    # Recency bonus (sold more recently = better)
    if comp.sold_date:
        try:
            sold_dt = datetime.strptime(comp.sold_date[:10], "%Y-%m-%d")
            days_ago = (datetime.now() - sold_dt).days
            if days_ago < 30:
                score += 0.05
            elif days_ago < 90:
                score += 0.02
            elif days_ago > 180:
                score -= 0.05
        except ValueError:
            pass

    return max(0.0, min(1.0, score))


# ── Adjustment engine ─────────────────────────────────────────────────

def _calculate_adjustments(subject: SubjectProperty, comp: CompProperty) -> dict:
    """Calculate dollar adjustments from comp to subject property."""
    adjustments = {}

    # Square footage adjustment
    if subject.sqft and comp.sqft:
        sqft_diff = subject.sqft - comp.sqft
        if sqft_diff != 0:
            adj = sqft_diff * ADJ_PER_SQFT
            adjustments["sqft"] = round(adj)

    # Bedroom adjustment
    if subject.bedrooms and comp.bedrooms:
        bed_diff = subject.bedrooms - comp.bedrooms
        if bed_diff != 0:
            adjustments["bedrooms"] = round(bed_diff * ADJ_PER_BEDROOM)

    # Bathroom adjustment
    if subject.bathrooms and comp.bathrooms:
        bath_diff = subject.bathrooms - comp.bathrooms
        if bath_diff != 0:
            adjustments["bathrooms"] = round(bath_diff * ADJ_PER_BATHROOM)

    # Age / year built adjustment
    if subject.year_built and comp.year_built:
        year_diff = subject.year_built - comp.year_built
        if year_diff != 0:
            adjustments["year_built"] = round(year_diff * ADJ_PER_YEAR_BUILT)

    # Lot size adjustment (capped)
    if subject.lot_sqft and comp.lot_sqft:
        lot_diff = subject.lot_sqft - comp.lot_sqft
        if lot_diff != 0:
            adj = lot_diff * ADJ_PER_LOT_SQFT
            adj = max(-ADJ_LOT_MAX, min(ADJ_LOT_MAX, adj))
            adjustments["lot_size"] = round(adj)

    # Garage adjustment
    garage_diff = subject.garage_spaces - comp.garage_spaces
    if garage_diff != 0:
        adjustments["garage"] = round(garage_diff * ADJ_PER_GARAGE)

    # Market conditions (time) adjustment
    if comp.sold_date:
        try:
            sold_dt = datetime.strptime(comp.sold_date[:10], "%Y-%m-%d")
            months_ago = (datetime.now() - sold_dt).days / 30.0
            if months_ago > 1:
                adj = comp.sold_price * MARKET_CONDITION_PCT_PER_MONTH * months_ago
                adjustments["market_conditions"] = round(adj)
        except ValueError:
            pass

    return adjustments


def _apply_adjustments(comp: CompProperty, adjustments: dict) -> float:
    """Apply adjustments to comp's sold price and return adjusted price."""
    total_adj = sum(adjustments.values())
    return comp.sold_price + total_adj


# ── Two-Bucket classification ────────────────────────────────────────

def _classify_bucket(comp: CompProperty) -> str:
    """Classify comp into Bucket A or Bucket B.

    Bucket A: Non-disclosure baseline comps — properties with limited price
              transparency (typical in TN as a non-disclosure state).
              Uses Zillow/MLS-reported data as proxy.
    Bucket B: Disclosure/verified comps — properties with confirmed sale
              prices from MLS (listed and sold through agent).

    In practice for TN (non-disclosure state), all comps come through
    Zillow/MLS so we classify based on data completeness:
    - Bucket B: Has complete MLS data (sold through agent, DOM tracked)
    - Bucket A: Limited data (off-market sale, FSBO, etc.)
    """
    if comp.days_on_market > 0:
        return "B"  # Was listed on MLS — has disclosure data
    return "A"  # No DOM data — likely off-market/non-disclosure sale


# ── ARV calculation ───────────────────────────────────────────────────

def calculate_arv(subject: SubjectProperty, comps: list[CompProperty]) -> ARVResult:
    """Calculate After Repair Value using Two-Bucket methodology.

    1. Score and rank comps by similarity
    2. Classify into Bucket A (non-disclosure) and Bucket B (disclosure)
    3. Apply property-specific adjustments
    4. Calculate weighted ARV with confidence bands
    """
    if not comps:
        return ARVResult(confidence="none", confidence_reason="No comparable sales found")

    # Score and sort by similarity
    for comp in comps:
        comp.similarity_score = _score_similarity(subject, comp)

    comps.sort(key=lambda c: c.similarity_score, reverse=True)

    # Take top comps
    selected = comps[:MAX_COMPS]

    # Classify buckets and calculate adjustments
    for comp in selected:
        comp.bucket = _classify_bucket(comp)
        comp.adjustments = _calculate_adjustments(subject, comp)
        comp.adjusted_price = _apply_adjustments(comp, comp.adjustments)

    bucket_a = [c for c in selected if c.bucket == "A"]
    bucket_b = [c for c in selected if c.bucket == "B"]

    # Calculate PPSF from adjusted prices
    ppsf_values = []
    for comp in selected:
        if comp.sqft:
            ppsf_values.append(comp.adjusted_price / comp.sqft)

    # Weighted ARV: Bucket B gets 70% weight, Bucket A gets 30%
    # (disclosure data is more reliable)
    adj_prices = [c.adjusted_price for c in selected if c.adjusted_price > 0]

    if not adj_prices:
        return ARVResult(confidence="none", confidence_reason="No valid adjusted prices")

    if bucket_b:
        bucket_b_avg = sum(c.adjusted_price for c in bucket_b) / len(bucket_b)
    else:
        bucket_b_avg = 0

    if bucket_a:
        bucket_a_avg = sum(c.adjusted_price for c in bucket_a) / len(bucket_a)
    else:
        bucket_a_avg = 0

    # Weighted average
    if bucket_b_avg and bucket_a_avg:
        arv_mid = bucket_b_avg * 0.70 + bucket_a_avg * 0.30
    elif bucket_b_avg:
        arv_mid = bucket_b_avg
    else:
        arv_mid = bucket_a_avg

    # Confidence bands based on comp spread
    spread = max(adj_prices) - min(adj_prices)
    spread_pct = (spread / arv_mid * 100) if arv_mid else 0

    # Conservative bands — low end is intentionally more conservative
    if spread_pct < 10:
        arv_low = arv_mid * 0.95
        arv_high = arv_mid * 1.05
        confidence = "high"
        confidence_reason = f"Tight comp spread ({spread_pct:.0f}%), {len(selected)} comps"
    elif spread_pct < 20:
        arv_low = arv_mid * 0.90
        arv_high = arv_mid * 1.08
        confidence = "medium"
        confidence_reason = f"Moderate comp spread ({spread_pct:.0f}%), {len(selected)} comps"
    else:
        arv_low = arv_mid * 0.85
        arv_high = arv_mid * 1.10
        confidence = "low"
        confidence_reason = f"Wide comp spread ({spread_pct:.0f}%) — verify with local knowledge"

    # Fewer comps = lower confidence
    if len(selected) < MIN_COMPS:
        confidence = "low"
        confidence_reason = f"Only {len(selected)} comps found (minimum {MIN_COMPS} recommended)"

    avg_adj = sum(abs(sum(c.adjustments.values())) for c in selected) / len(selected) if selected else 0

    return ARVResult(
        arv_low=round(arv_low),
        arv_mid=round(arv_mid),
        arv_high=round(arv_high),
        confidence=confidence,
        confidence_reason=confidence_reason,
        ppsf_avg=round(sum(ppsf_values) / len(ppsf_values), 2) if ppsf_values else 0,
        ppsf_range=(round(min(ppsf_values), 2), round(max(ppsf_values), 2)) if ppsf_values else (0, 0),
        comp_count=len(selected),
        bucket_a_count=len(bucket_a),
        bucket_b_count=len(bucket_b),
        avg_adjustment=round(avg_adj),
        spread_pct=round(spread_pct, 1),
    )


# ── Excel report generation ──────────────────────────────────────────

# Styles
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_MONEY_FMT = '#,##0'
_PCT_FMT = '0.0%'
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_VALUE_FONT = Font(name="Calibri", bold=True, size=13, color="222222")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT = Font(name="Calibri", bold=True, color="006100")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_column_widths(ws, min_width: int = 12, max_width: int = 35) -> None:
    """Auto-size columns based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _fmt_money(value: float) -> str:
    """Format number as currency string."""
    if not value:
        return "$0"
    return f"${value:,.0f}"


def generate_comp_report(subject: SubjectProperty, comps: list[CompProperty],
                         arv: ARVResult, output_path: str) -> str:
    """Generate a 7-tab Excel workbook comp report.

    Tabs:
    1. Executive Summary — subject property + ARV range
    2. Subject Property — full detail
    3. Comparable Sales — all comps with distance, date, similarity
    4. Adjustments Detail — per-comp adjustment breakdown
    5. Market Analysis — PPSF trends, DOM, market direction
    6. ARV Calculation — Two-Bucket weighted result with confidence bands
    7. Sources & Notes
    """
    wb = Workbook()

    # ── Tab 1: Executive Summary ──────────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"

    ws.cell(row=1, column=1, value="Comp Analysis Report").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"{subject.address}, {subject.city}, {subject.state} {subject.zip_code}").font = _SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = _LABEL_FONT

    row = 5
    summary_data = [
        ("ARV (Low)", _fmt_money(arv.arv_low)),
        ("ARV (Mid — Recommended)", _fmt_money(arv.arv_mid)),
        ("ARV (High)", _fmt_money(arv.arv_high)),
        ("", ""),
        ("Confidence Level", arv.confidence.upper()),
        ("Confidence Reason", arv.confidence_reason),
        ("", ""),
        ("Avg PPSF", f"${arv.ppsf_avg:,.2f}"),
        ("PPSF Range", f"${arv.ppsf_range[0]:,.2f} — ${arv.ppsf_range[1]:,.2f}"),
        ("Comps Analyzed", str(arv.comp_count)),
        ("Bucket A (Non-Disclosure)", str(arv.bucket_a_count)),
        ("Bucket B (Disclosure/MLS)", str(arv.bucket_b_count)),
        ("Avg Gross Adjustment", _fmt_money(arv.avg_adjustment)),
        ("Comp Spread", f"{arv.spread_pct:.1f}%"),
        ("", ""),
        ("Subject Zestimate", _fmt_money(subject.zestimate)),
        ("Subject Property Type", subject.property_type),
        ("Subject Sqft", f"{subject.sqft:,}" if subject.sqft else "N/A"),
        ("Subject Bed/Bath", f"{subject.bedrooms}bd / {subject.bathrooms}ba"),
        ("Subject Year Built", str(subject.year_built) if subject.year_built else "N/A"),
    ]
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = _VALUE_FONT
        if label == "Confidence Level":
            if arv.confidence == "high":
                cell.fill = _GREEN_FILL
                cell.font = _GREEN_FONT
            elif arv.confidence == "medium":
                cell.fill = _YELLOW_FILL
            else:
                cell.fill = _RED_FILL
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35

    # ── Tab 2: Subject Property ──────────────────────────────────────
    ws2 = wb.create_sheet("Subject Property")
    ws2.cell(row=1, column=1, value="Subject Property Details").font = _TITLE_FONT

    props = [
        ("Address", subject.address),
        ("City", subject.city),
        ("State", subject.state),
        ("ZIP", subject.zip_code),
        ("Latitude", str(subject.latitude) if subject.latitude else ""),
        ("Longitude", str(subject.longitude) if subject.longitude else ""),
        ("Property Type", subject.property_type),
        ("Square Feet", f"{subject.sqft:,}" if subject.sqft else ""),
        ("Bedrooms", str(subject.bedrooms)),
        ("Bathrooms", str(subject.bathrooms)),
        ("Year Built", str(subject.year_built) if subject.year_built else ""),
        ("Lot Size (sqft)", f"{subject.lot_sqft:,}" if subject.lot_sqft else ""),
        ("Garage Spaces", str(subject.garage_spaces)),
        ("Zestimate", _fmt_money(subject.zestimate)),
        ("MLS Status", subject.mls_status),
        ("Last Sold Date", subject.last_sold_date),
        ("Last Sold Price", _fmt_money(subject.last_sold_price)),
    ]
    for i, (label, value) in enumerate(props, 3):
        ws2.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws2.cell(row=i, column=2, value=value).font = _VALUE_FONT
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 30

    # ── Tab 3: Comparable Sales ──────────────────────────────────────
    ws3 = wb.create_sheet("Comparable Sales")
    ws3.cell(row=1, column=1, value="Comparable Sales").font = _TITLE_FONT

    comp_headers = ["#", "Address", "City", "ZIP", "Distance (mi)", "Sold Price",
                    "Sold Date", "Sqft", "Bed", "Bath", "Year Built", "PPSF",
                    "Similarity", "Bucket", "Adjusted Price"]
    _write_header_row(ws3, 3, comp_headers)

    for i, comp in enumerate(comps[:MAX_COMPS], 1):
        row = i + 3
        values = [
            i, comp.address, comp.city, comp.zip_code,
            comp.distance_miles, comp.sold_price, comp.sold_date,
            comp.sqft, comp.bedrooms, comp.bathrooms, comp.year_built,
            round(comp.ppsf, 2), f"{comp.similarity_score:.0%}",
            comp.bucket, comp.adjusted_price,
        ]
        for col, val in enumerate(values, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            if col in (6, 15):  # Money columns
                cell.number_format = _MONEY_FMT
            cell.border = _THIN_BORDER

    _auto_column_widths(ws3)

    # ── Tab 4: Adjustments Detail ────────────────────────────────────
    ws4 = wb.create_sheet("Adjustments Detail")
    ws4.cell(row=1, column=1, value="Per-Comp Adjustment Breakdown").font = _TITLE_FONT

    adj_types = ["sqft", "bedrooms", "bathrooms", "year_built", "lot_size", "garage", "market_conditions"]
    adj_headers = ["Comp #", "Address", "Sold Price"] + [a.replace("_", " ").title() for a in adj_types] + ["Total Adj", "Adjusted Price"]
    _write_header_row(ws4, 3, adj_headers)

    for i, comp in enumerate(comps[:MAX_COMPS], 1):
        row = i + 3
        values = [i, comp.address, comp.sold_price]
        total_adj = 0
        for adj_type in adj_types:
            adj_val = comp.adjustments.get(adj_type, 0)
            total_adj += adj_val
            values.append(adj_val)
        values.append(total_adj)
        values.append(comp.adjusted_price)
        for col, val in enumerate(values, 1):
            cell = ws4.cell(row=row, column=col, value=val)
            if col >= 3:
                cell.number_format = _MONEY_FMT
            cell.border = _THIN_BORDER

    _auto_column_widths(ws4)

    # ── Tab 5: Market Analysis ───────────────────────────────────────
    ws5 = wb.create_sheet("Market Analysis")
    ws5.cell(row=1, column=1, value="Market Analysis").font = _TITLE_FONT

    # PPSF analysis
    ws5.cell(row=3, column=1, value="Price Per Square Foot Analysis").font = _SUBTITLE_FONT
    ppsf_data = [
        ("Average PPSF", f"${arv.ppsf_avg:,.2f}"),
        ("PPSF Range", f"${arv.ppsf_range[0]:,.2f} — ${arv.ppsf_range[1]:,.2f}"),
        ("Subject Implied Value (Avg PPSF)", _fmt_money(arv.ppsf_avg * subject.sqft) if subject.sqft else "N/A"),
    ]
    for i, (label, value) in enumerate(ppsf_data, 4):
        ws5.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws5.cell(row=i, column=2, value=value).font = _VALUE_FONT

    # Days on market
    ws5.cell(row=8, column=1, value="Days on Market Analysis").font = _SUBTITLE_FONT
    dom_values = [c.days_on_market for c in comps[:MAX_COMPS] if c.days_on_market > 0]
    if dom_values:
        ws5.cell(row=9, column=1, value="Average DOM").font = _LABEL_FONT
        ws5.cell(row=9, column=2, value=f"{sum(dom_values) / len(dom_values):.0f} days").font = _VALUE_FONT
        ws5.cell(row=10, column=1, value="Median DOM").font = _LABEL_FONT
        sorted_dom = sorted(dom_values)
        median_dom = sorted_dom[len(sorted_dom) // 2]
        ws5.cell(row=10, column=2, value=f"{median_dom} days").font = _VALUE_FONT

    # Market direction
    ws5.cell(row=12, column=1, value="Market Direction").font = _SUBTITLE_FONT
    ws5.cell(row=13, column=1, value="Monthly Appreciation Rate").font = _LABEL_FONT
    ws5.cell(row=13, column=2, value=f"{MARKET_CONDITION_PCT_PER_MONTH * 100:.1f}%").font = _VALUE_FONT
    ws5.cell(row=14, column=1, value="Annualized Appreciation").font = _LABEL_FONT
    ws5.cell(row=14, column=2, value=f"{MARKET_CONDITION_PCT_PER_MONTH * 12 * 100:.1f}%").font = _VALUE_FONT

    ws5.column_dimensions["A"].width = 35
    ws5.column_dimensions["B"].width = 25

    # ── Tab 6: ARV Calculation ───────────────────────────────────────
    ws6 = wb.create_sheet("ARV Calculation")
    ws6.cell(row=1, column=1, value="Two-Bucket ARV Calculation").font = _TITLE_FONT

    ws6.cell(row=3, column=1, value="Methodology").font = _SUBTITLE_FONT
    ws6.cell(row=4, column=1, value="Bucket A (Non-Disclosure): 30% weight — Off-market/FSBO sales with limited price transparency").font = _LABEL_FONT
    ws6.cell(row=5, column=1, value="Bucket B (Disclosure/MLS): 70% weight — Agent-listed sales with confirmed pricing").font = _LABEL_FONT
    ws6.cell(row=6, column=1, value="Tennessee is a non-disclosure state. All data sourced via Zillow/MLS.").font = _LABEL_FONT

    ws6.cell(row=8, column=1, value="Bucket A Comps").font = _SUBTITLE_FONT
    bucket_a = [c for c in comps[:MAX_COMPS] if c.bucket == "A"]
    bucket_b = [c for c in comps[:MAX_COMPS] if c.bucket == "B"]
    if bucket_a:
        avg_a = sum(c.adjusted_price for c in bucket_a) / len(bucket_a)
        ws6.cell(row=9, column=1, value=f"Count: {len(bucket_a)}  |  Avg Adjusted: {_fmt_money(avg_a)}").font = _VALUE_FONT
    else:
        ws6.cell(row=9, column=1, value="No Bucket A comps").font = _LABEL_FONT

    ws6.cell(row=11, column=1, value="Bucket B Comps").font = _SUBTITLE_FONT
    if bucket_b:
        avg_b = sum(c.adjusted_price for c in bucket_b) / len(bucket_b)
        ws6.cell(row=12, column=1, value=f"Count: {len(bucket_b)}  |  Avg Adjusted: {_fmt_money(avg_b)}").font = _VALUE_FONT
    else:
        ws6.cell(row=12, column=1, value="No Bucket B comps").font = _LABEL_FONT

    ws6.cell(row=14, column=1, value="Final ARV").font = _SUBTITLE_FONT
    arv_display = [
        ("ARV Low (Conservative)", _fmt_money(arv.arv_low)),
        ("ARV Mid (Recommended)", _fmt_money(arv.arv_mid)),
        ("ARV High (Optimistic)", _fmt_money(arv.arv_high)),
        ("Confidence", arv.confidence.upper()),
        ("Reason", arv.confidence_reason),
    ]
    for i, (label, value) in enumerate(arv_display, 15):
        ws6.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws6.cell(row=i, column=2, value=value).font = _VALUE_FONT

    ws6.column_dimensions["A"].width = 40
    ws6.column_dimensions["B"].width = 30

    # ── Tab 7: Sources & Notes ───────────────────────────────────────
    ws7 = wb.create_sheet("Sources & Notes")
    ws7.cell(row=1, column=1, value="Sources & Notes").font = _TITLE_FONT

    notes = [
        "Data Source: OpenWeb Ninja Real-Time Zillow Data API",
        "Comparable sales sourced from Zillow's similar-sale-homes endpoint",
        "",
        "Adjustment Methodology:",
        f"  Square Footage: ${ADJ_PER_SQFT:,.0f} per sqft difference",
        f"  Bedrooms: ${ADJ_PER_BEDROOM:,.0f} per bedroom difference",
        f"  Bathrooms: ${ADJ_PER_BATHROOM:,.0f} per bathroom difference",
        f"  Year Built: ${ADJ_PER_YEAR_BUILT:,.0f} per year difference",
        f"  Lot Size: ${ADJ_PER_LOT_SQFT:,.2f} per sqft (capped at ${ADJ_LOT_MAX:,.0f})",
        f"  Garage: ${ADJ_PER_GARAGE:,.0f} per stall difference",
        f"  Market Conditions: {MARKET_CONDITION_PCT_PER_MONTH * 100:.1f}% per month appreciation",
        "",
        "Two-Bucket Weighting:",
        "  Bucket A (Non-Disclosure): 30% weight",
        "  Bucket B (Disclosure/MLS): 70% weight",
        "",
        "Confidence Bands:",
        "  High (<10% spread): ±5% of mid ARV",
        "  Medium (10-20% spread): -10%/+8% of mid ARV",
        "  Low (>20% spread): -15%/+10% of mid ARV",
        "",
        "Conservative bias: low-end ARV is intentionally wider.",
        "A high ARV that doesn't hold up kills your deal.",
        "A conservative ARV that comes in low leaves room for upside.",
        "",
        f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Region: Knoxville / East Tennessee",
    ]
    for i, note in enumerate(notes, 3):
        ws7.cell(row=i, column=1, value=note).font = _LABEL_FONT

    ws7.column_dimensions["A"].width = 70

    # Save
    wb.save(output_path)
    logger.info("Comp report saved to %s", output_path)
    return output_path


# ── Main entry point ──────────────────────────────────────────────────

def run_comp_analysis(address: str, city: str = "", state: str = "",
                      zip_code: str = "", radius: float = DEFAULT_RADIUS_MILES,
                      months: int = DEFAULT_MONTHS_BACK,
                      output_path: str = "") -> dict:
    """Run a full comp analysis for a property and generate the report.

    Returns a dict with ARV results and the output file path.
    """
    logger.info("Starting comp analysis for: %s %s %s %s", address, city, state, zip_code)

    # Step 1: Fetch subject property details
    subject = fetch_subject_property(address, city, state, zip_code)
    if not subject:
        logger.error("Could not fetch subject property data")
        return {"error": "Could not fetch subject property data"}

    logger.info("Subject: %s — %s sqft, %dbd/%sba, built %s, Zestimate %s",
                subject.address, f"{subject.sqft:,}" if subject.sqft else "?",
                subject.bedrooms, subject.bathrooms,
                subject.year_built or "?", _fmt_money(subject.zestimate))

    # Step 2: Fetch comparable sales
    comps = fetch_comparable_sales(subject, radius, months)
    if not comps:
        logger.warning("No comparable sales found — try expanding radius or time window")
        return {"error": "No comparable sales found", "subject": subject}

    # Step 3: Calculate ARV
    arv = calculate_arv(subject, comps)
    logger.info("ARV: %s (low) / %s (mid) / %s (high) — %s confidence",
                _fmt_money(arv.arv_low), _fmt_money(arv.arv_mid),
                _fmt_money(arv.arv_high), arv.confidence)

    # Step 4: Generate report
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_address = "".join(c if c.isalnum() or c in " -" else "_" for c in address)[:40]
        output_path = str(config.OUTPUT_DIR / f"comp_report_{safe_address}_{timestamp}.xlsx")

    report_path = generate_comp_report(subject, comps, arv, output_path)

    return {
        "subject": subject,
        "comps": comps[:MAX_COMPS],
        "arv": arv,
        "report_path": report_path,
    }
