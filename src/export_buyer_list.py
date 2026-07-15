"""Top 25 Wholesale-Friendly Cash Buyer List Generator.

Filters the nationwide buyers database for a target county, excludes
institutional/government buyers, scores for wholesale-friendliness,
resolves decision-makers behind entities, and outputs a polished Excel.

Usage:
  python src/export_buyer_list.py --county Knox --state TN --top 25 --research
"""

import argparse
import csv
import json
import logging
import random
import re
import time
from datetime import datetime

import anthropic
from ddgs import DDGS
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config

logger = logging.getLogger(__name__)

# ── Paths ────────────────────────────────────────────────────────────────

NATIONWIDE_CSV = (
    config.PROJECT_ROOT
    / "Skills for REI"
    / "extracted"
    / "skills_expanded"
    / "buyer-prospector"
    / "buyer-prospector"
    / "data"
    / "nationwide_buyers.csv"
)

# ── Institutional Exclusion ──────────────────────────────────────────────

HARD_EXCLUSIONS = {
    "OPENDOOR PROPERTY TRUST",
    "ARMM ASSET COMPANY 2 LLC",
    "SOLID GROUND SALES LLC",
    "DR HORTON INC",
    "CLAYTON PROPERTIES GROUP INC",
    "TENNESSEE HOUSING DEVELOPMENT AGENCY",
}

IBUYER_KEYWORDS = [
    "OPENDOOR", "OFFERPAD", "ZILLOW", "REDFIN", "KNOCK",
    "INVITATION HOMES", "PROGRESS RESIDENTIAL", "AMERICAN HOMES 4 RENT",
    "TRICON", "PRETIUM", "CERBERUS", "BLACKSTONE", "COLONY", "STARWOOD",
    "FANNIE MAE", "FREDDIE MAC",
]

GOVERNMENT_KEYWORDS = [
    "HOUSING AUTHORITY", "DEVELOPMENT AGENCY", "CITY OF", "COUNTY OF",
    "STATE OF", "HOUSING DEVELOPMENT", "HUD",
]

VOLUME_THRESHOLD = 100  # Auto-exclude if total purchases >= this

# ── Entity Classification ────────────────────────────────────────────────

def categorize_entity(name):
    """Classify a buyer name into entity types."""
    if not name:
        return "UNKNOWN"
    upper = str(name).upper().strip()

    if "LLC" in upper or "L.L.C" in upper:
        return "LLC"
    if "TRUST" in upper or upper.endswith(" TRU"):
        return "TRUST"
    if any(kw in upper for kw in [" INC", " INC.", "INCORPORATED"]):
        return "CORPORATION"
    if any(kw in upper for kw in [" CORP", " CORP.", "CORPORATION"]):
        return "CORPORATION"
    if upper.endswith("CORPORATIO"):
        return "CORPORATION"
    if "ESTATE OF" in upper or "ESTATE" in upper:
        return "ESTATE"
    if any(kw in upper for kw in [" LP", " LLP", "LIMITED PARTNERSHIP"]):
        return "LIMITED PARTNERSHIP"
    if any(kw in upper for kw in ["COUNTY", "CITY OF", "STATE OF",
                                   "HOUSING AUTHORITY", "GOVERNMENT"]):
        return "GOVERNMENT/AGENCY"

    business_keywords = [
        "PROPERTIES", "HOLDINGS", "INVESTMENTS", "CAPITAL", "GROUP",
        "PARTNERS", "COMPANY", "ENTERPRISES", "VENTURES", "REALTY",
        "HOMES", "REAL ESTATE", "BUILDERS", "CONSTRUCTION", "MANAGEMENT",
        "SOLUTIONS", "SERVICES", "ASSOCIATES", "FUNDING", "ACQUISITIONS",
        "BUYERS", "RENOVATIONS", "DEVELOPMENT", "CONSULTING",
    ]
    words = name.split()
    if len(words) <= 3 and not any(kw in upper for kw in business_keywords):
        return "INDIVIDUAL"
    if any(kw in upper for kw in business_keywords):
        return "OTHER ENTITY"
    if len(words) > 3:
        return "OTHER ENTITY"
    return "INDIVIDUAL"


# ── Data Loading & Filtering ─────────────────────────────────────────────

def load_and_filter(county: str, state: str, min_purchases: int = 2) -> list[dict]:
    """Load nationwide CSV, filter to target county+state, deduplicate."""
    logger.info("Loading nationwide buyers from %s", NATIONWIDE_CSV)

    rows = []
    with open(NATIONWIDE_CSV, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            cn = (row.get("County Name") or "").strip().lower()
            cs = (row.get("County State") or "").strip().upper()
            if cn == county.lower() and cs == state.upper():
                row["BuyerPurchases6MSum"] = int(row.get("BuyerPurchases6MSum", 0))
                rows.append(row)

    logger.info("Found %d raw records for %s County, %s", len(rows), county, state)

    # Deduplicate by BuyerFullName — sum purchases, keep first address
    # Also merge spelling variants (VASQUEZ/VAZQUEZ, etc.)
    seen = {}
    for row in rows:
        name = (row.get("BuyerFullName") or "").strip().upper()
        # Normalize common spelling variants for dedup key
        dedup_key = name.replace("Z", "S").replace("PH", "F")
        # Check both exact name and normalized variant
        matched_key = None
        if name in seen:
            matched_key = name
        elif dedup_key in {k.replace("Z", "S").replace("PH", "F") for k in seen}:
            # Find the existing key this matches
            for existing_key in seen:
                if existing_key.replace("Z", "S").replace("PH", "F") == dedup_key:
                    matched_key = existing_key
                    break

        if matched_key:
            seen[matched_key]["BuyerPurchases6MSum"] += row["BuyerPurchases6MSum"]
        else:
            seen[name] = dict(row)
            seen[name]["BuyerFullName_original"] = row.get("BuyerFullName", "").strip()

    deduped = list(seen.values())
    # Apply minimum purchase filter
    deduped = [r for r in deduped if r["BuyerPurchases6MSum"] >= min_purchases]
    deduped.sort(key=lambda r: r["BuyerPurchases6MSum"], reverse=True)

    logger.info("After dedup + min %d purchases: %d unique buyers",
                min_purchases, len(deduped))
    return deduped


def exclude_institutional(buyers: list[dict]) -> tuple[list[dict], list[dict]]:
    """Remove institutional, government, and iBuyer entities.

    Returns (kept, excluded) where each excluded entry has an 'exclusion_reason'.
    """
    kept = []
    excluded = []

    for buyer in buyers:
        name = (buyer.get("BuyerFullName") or "").strip().upper()
        original_name = buyer.get("BuyerFullName_original", name)
        purchases = buyer["BuyerPurchases6MSum"]
        reason = None

        # Hard exclusion list
        if name in {n.upper() for n in HARD_EXCLUSIONS}:
            reason = "Hard exclusion list"

        # Volume threshold
        elif purchases >= VOLUME_THRESHOLD:
            reason = f"Institutional volume ({purchases} purchases)"

        # iBuyer/institutional patterns
        elif any(kw in name for kw in IBUYER_KEYWORDS):
            reason = "iBuyer/institutional pattern"

        # Government keywords
        elif any(kw in name for kw in GOVERNMENT_KEYWORDS):
            reason = "Government/agency"

        # Single-asset institutional pattern: "[WORD] OWNER LLC"
        elif re.search(r"\bOWNER\s+LLC\b", name):
            reason = "Single-asset institutional entity"

        # HUD-related entities
        elif "HHI-HUD" in name or "HUD " in name:
            reason = "HUD-related institutional"

        if reason:
            buyer["exclusion_reason"] = reason
            excluded.append(buyer)
            logger.debug("Excluded: %s — %s", original_name, reason)
        else:
            kept.append(buyer)

    logger.info("Kept %d buyers, excluded %d", len(kept), len(excluded))
    return kept, excluded


# ── Wholesale-Friendliness Scoring ───────────────────────────────────────

KNOX_ZIPS = {
    "37901", "37902", "37909", "37912", "37914", "37915", "37916",
    "37917", "37918", "37919", "37920", "37921", "37922", "37923",
    "37924", "37927", "37928", "37929", "37930", "37931", "37932",
    "37933", "37934", "37938", "37939", "37940", "37950",
}

# Adjacent TN county ZIPs (Blount, Anderson, Union, Loudon, Roane, etc.)
ADJACENT_ZIPS_PREFIX = {"377", "378", "378"}  # rough prefix match for nearby


def _score_local(buyer: dict) -> float:
    """Score 0-30: local presence."""
    buyer_state = (buyer.get("BuyerState") or buyer.get("County State") or "").strip().upper()
    buyer_zip = (buyer.get("BuyerZIP") or "")[:5]
    buyer_city = (buyer.get("BuyerCity") or "").strip().upper()

    # Knox County ZIP
    if buyer_zip in KNOX_ZIPS:
        return 30.0
    # Knoxville city (catches ZIPs we might have missed)
    if buyer_city == "KNOXVILLE" or buyer_city == "FARRAGUT" or buyer_city == "POWELL":
        return 30.0
    # Adjacent TN counties
    if buyer_state == "TN" and buyer_zip[:3] in {"377", "378", "376", "373"}:
        return 20.0
    # Other TN
    if buyer_state == "TN":
        return 15.0
    # Out of state
    return 5.0


def _score_activity(purchases: int) -> float:
    """Score 0-25: activity sweet spot (6-12 is ideal for wholesale)."""
    if purchases <= 1:
        return 5.0
    if purchases == 2:
        return 10.0
    if 3 <= purchases <= 5:
        return 18.0
    if 6 <= purchases <= 12:
        return 25.0
    if 13 <= purchases <= 20:
        return 20.0
    if 21 <= purchases <= 30:
        return 12.0
    if 31 <= purchases <= 50:
        return 5.0
    return 2.0  # >50


def _score_entity_type(entity_type: str) -> float:
    """Score 0-20: entity type fit for wholesaling."""
    scores = {
        "LLC": 20.0,
        "INDIVIDUAL": 18.0,
        "TRUST": 10.0,
        "CORPORATION": 8.0,
        "ESTATE": 6.0,
        "LIMITED PARTNERSHIP": 8.0,
        "OTHER ENTITY": 12.0,
        "GOVERNMENT/AGENCY": 0.0,
        "UNKNOWN": 5.0,
    }
    return scores.get(entity_type, 5.0)


INVESTOR_KEYWORDS = [
    "PROPERTIES", "HOMES", "INVESTMENTS", "CAPITAL", "HOLDINGS",
    "VENTURES", "REALTY", "REAL ESTATE", "FLIPS", "BUYERS", "ACQUISITIONS",
    "RENOVATIONS", "SOLUTIONS", "HOME",
]


def _score_name(name: str) -> float:
    """Score 0-15: name signals investor activity."""
    upper = name.upper()
    if any(kw in upper for kw in INVESTOR_KEYWORDS):
        return 15.0
    # Generic business name
    if any(kw in upper for kw in ["BUILDERS", "CONSTRUCTION", "MANAGEMENT",
                                   "SERVICES", "CONSULTING", "GROUP"]):
        return 10.0
    # Acronym-only LLC (HBB LLC, RHBTN LLC)
    words = upper.replace("LLC", "").replace("INC", "").strip().split()
    if len(words) == 1 and len(words[0]) <= 5:
        return 8.0
    return 10.0  # Default for individuals and normal names


def _score_address(buyer: dict) -> float:
    """Score 0-10: address quality."""
    addr = (buyer.get("BuyerAddress") or "").strip()
    if not addr:
        return 0.0
    if addr.upper().startswith("PO BOX") or addr.upper().startswith("P.O."):
        return 5.0
    return 10.0


def score_buyers(buyers: list[dict]) -> list[dict]:
    """Score and rank buyers for wholesale-friendliness."""
    for buyer in buyers:
        name = buyer.get("BuyerFullName_original") or buyer.get("BuyerFullName", "")
        entity_type = buyer.get("EntityType", "UNKNOWN")
        purchases = buyer["BuyerPurchases6MSum"]

        local = _score_local(buyer)
        activity = _score_activity(purchases)
        entity = _score_entity_type(entity_type)
        name_score = _score_name(name)
        address = _score_address(buyer)

        total = local + activity + entity + name_score + address

        buyer["score_total"] = round(total, 1)
        buyer["score_local"] = round(local, 1)
        buyer["score_activity"] = round(activity, 1)
        buyer["score_entity"] = round(entity, 1)
        buyer["score_name"] = round(name_score, 1)
        buyer["score_address"] = round(address, 1)

    buyers.sort(key=lambda b: b["score_total"], reverse=True)
    for i, buyer in enumerate(buyers):
        buyer["rank"] = i + 1

    return buyers


# ── Decision-Maker Resolution ────────────────────────────────────────────

_LLC_RE = re.compile(r"\b(?:LLC|L\.L\.C)\b", re.IGNORECASE)
_CORP_RE = re.compile(r"\b(?:INC|INCORPORATED|CORP|CORPORATION)\b", re.IGNORECASE)

SEARCH_DELAY_MIN = 0.5
SEARCH_DELAY_MAX = 1.0
LLM_MODEL = "claude-haiku-4-5-20251001"
LLM_MAX_TOKENS = 256

ENTITY_SYSTEM_PROMPT = (
    "You are a business entity research assistant. You analyze search results "
    "to identify the real person behind a business entity. Return only valid JSON."
)

ENTITY_EXTRACT_PROMPT = """You are analyzing search results to find the real person behind a business entity.

Entity: "{entity_name}" (type: {entity_type})

Search results:
{snippets}

Extract the following if found:
- person_name: Full name of a person associated with this entity (first and last name)
- role: Their role (registered_agent, member, manager, officer, trustee, partner, principal, organizer)
- confidence: high (exact match from official records or business listing), medium (likely match from directory), low (name mentioned but unclear relationship)

If multiple people are found, return the one most likely to be the decision-maker (owner/member > registered agent > officer).

Return JSON: {{"person_name": "...", "role": "...", "confidence": "..."}}
If no person can be identified, return: {{"person_name": "", "role": "", "confidence": ""}}"""


def _parse_individual_name(name: str) -> dict:
    """Parse deed-format individual name (LASTNAME FIRSTNAME MIDDLE) into parts."""
    parts = name.strip().split()
    if len(parts) >= 2:
        # Deed format: WARD J CALVIN → first="J Calvin", last="Ward"
        last = parts[0].title()
        first = " ".join(p.title() for p in parts[1:])
        full = f"{first} {last}"
        return {
            "dm_full": full,
            "dm_first": first,
            "dm_last": last,
            "dm_role": "Individual",
            "dm_confidence": "high",
            "dm_source": "Direct (individual buyer)",
        }
    return {
        "dm_full": name.title(),
        "dm_first": "",
        "dm_last": name.title(),
        "dm_role": "Individual",
        "dm_confidence": "high",
        "dm_source": "Direct (individual buyer)",
    }


def _try_parse_entity_name(name: str, entity_type: str) -> dict | None:
    """Try to extract a person name from the entity name (free fast path)."""
    if not name:
        return None

    upper = name.upper().strip()

    # Trust: extract grantor/trustee from trust name
    if entity_type == "TRUST" or upper.endswith(" TRU"):
        m = config.TRUST_NAME_RE.match(name)
        if m:
            extracted = m.group(1).strip()
            parts = extracted.split()
            business_words = {
                "FIRST", "NATIONAL", "AMERICAN", "COMMUNITY", "BANK",
                "FINANCIAL", "INVESTMENT", "CAPITAL", "HOLDINGS",
                "PROPERTIES", "MANAGEMENT", "GROUP", "SERVICES",
            }
            # Reject "Family" trust names — not a real person
            non_person = {"FAMILY", "COMMUNITY", "PROPERTY"}
            if len(parts) >= 2 and not any(w.upper() in business_words for w in parts) \
                    and not any(w.upper() in non_person for w in parts):
                full = extracted.title()
                first = " ".join(p.title() for p in parts[:-1])
                last = parts[-1].title()
                return {
                    "dm_full": full,
                    "dm_first": first,
                    "dm_last": last,
                    "dm_role": "Trustee",
                    "dm_confidence": "high",
                    "dm_source": "Trust name parsing",
                }
            if len(parts) == 1 and parts[0].upper() not in business_words:
                return {
                    "dm_full": parts[0].title(),
                    "dm_first": "",
                    "dm_last": parts[0].title(),
                    "dm_role": "Trustee",
                    "dm_confidence": "medium",
                    "dm_source": "Trust name parsing (surname only)",
                }
        return None

    # Estate: extract decedent name
    if entity_type == "ESTATE":
        m = config.ESTATE_OF_RE.match(name)
        if m:
            extracted = m.group(1).strip()
            parts = extracted.split()
            if parts:
                full = extracted.title()
                first = " ".join(p.title() for p in parts[:-1]) if len(parts) > 1 else ""
                last = parts[-1].title()
                return {
                    "dm_full": full,
                    "dm_first": first,
                    "dm_last": last,
                    "dm_role": "Executor",
                    "dm_confidence": "high",
                    "dm_source": "Estate name parsing",
                }
        return None

    # LLC with personal surname: "JOHNSON PROPERTIES LLC" → "Johnson"
    if entity_type == "LLC":
        cleaned = _LLC_RE.sub("", name).strip()
        cleaned = re.sub(r"[,.]", "", cleaned).strip()
        words = cleaned.split()
        generic = {
            "PROPERTIES", "HOLDINGS", "INVESTMENTS", "CAPITAL", "GROUP",
            "ENTERPRISES", "VENTURES", "REALTY", "HOMES", "REAL",
            "BUILDERS", "CONSTRUCTION", "MANAGEMENT", "SOLUTIONS",
            "SERVICES", "ASSOCIATES", "FUNDING", "ACQUISITIONS",
            "BUYERS", "RENOVATIONS", "DEVELOPMENT", "CONSULTING",
        }
        non_names = {
            "FIRST", "SECOND", "THIRD", "BEST", "PRIME", "TOP",
            "QUICK", "FAST", "SMART", "GOOD", "GREAT", "FAIR",
            "NEW", "OLD", "BIG", "LITTLE", "GLOBAL", "NATIONAL",
            "AMERICAN", "SOUTHERN", "EASTERN", "WESTERN", "NORTHERN",
            "CENTRAL", "PACIFIC", "ATLANTIC", "MOUNTAIN", "VALLEY",
            "LAKE", "RIVER", "HILL", "SUMMIT", "PEAK", "EAST",
            "BALL", "APEX", "GOD", "MISS", "HOME", "STORE",
            "DEFINITY", "MEDALLION", "ATHENA", "STOREHOUSE", "TUFF",
            "DHARMA", "ELEVENTH", "REBUILT", "ELEVATE",
            "RIVERWOOD", "STANLEY", "MESANA", "BLUE", "HARVEST",
            "MARTIN", "GATEWAY", "STOREHOUSE", "SELL",
            # Common non-name LLC prefixes
            "MRA", "GDP", "BJT", "HBB", "RHBTN", "DEMN",
            "VOLHOMES", "SMITHBILT", "SASKCUS",
        }
        if len(words) == 2 and words[1].upper() in generic:
            candidate = words[0]
            if candidate.upper() not in non_names and len(candidate) >= 3:
                return {
                    "dm_full": candidate.title(),
                    "dm_first": "",
                    "dm_last": candidate.title(),
                    "dm_role": "Member (probable surname)",
                    "dm_confidence": "low",
                    "dm_source": "LLC name parsing (surname guess)",
                }

    return None


def _search_entity(entity_name: str, state: str = "Tennessee") -> list[dict]:
    """Search DuckDuckGo for entity registration info."""
    query = f'"{entity_name}" {state} registered agent OR member OR officer'
    try:
        results = DDGS().text(query, max_results=8)
    except Exception as e:
        logger.debug("Search failed for '%s': %s", query, e)
        return []

    filtered = []
    for r in results:
        url = r.get("href", "")
        title = r.get("title", "")
        snippet = r.get("body", "")
        if url and (title or snippet):
            filtered.append({"url": url, "title": title, "snippet": snippet})
    return filtered


def _parse_with_llm(entity_name: str, entity_type: str,
                    search_results: list[dict], api_key: str) -> dict | None:
    """Use Claude Haiku to extract person info from search results."""
    if not search_results or not api_key:
        return None

    snippets = []
    for r in search_results[:6]:
        title = r.get("title", "")
        snippet = r.get("snippet", "")
        url = r.get("url", "")
        snippets.append(f"[{title}] ({url})\n{snippet}")

    combined = "\n\n".join(snippets)
    if len(combined) > 4000:
        combined = combined[:4000]

    prompt = ENTITY_EXTRACT_PROMPT.format(
        entity_name=entity_name,
        entity_type=entity_type,
        snippets=combined,
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model=LLM_MODEL,
            max_tokens=LLM_MAX_TOKENS,
            system=ENTITY_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": prompt}],
        )
        result_text = response.content[0].text.strip()
        result_text = re.sub(r"^```(?:json)?\s*", "", result_text)
        result_text = re.sub(r"\s*```$", "", result_text)
        brace_match = re.search(r"\{[^{}]*\}", result_text)
        if brace_match:
            result_text = brace_match.group(0)
        parsed = json.loads(result_text)

        person_name = parsed.get("person_name", "").strip()
        if person_name:
            parts = person_name.split()
            first = " ".join(parts[:-1]) if len(parts) > 1 else ""
            last = parts[-1] if parts else ""
            return {
                "dm_full": person_name,
                "dm_first": first,
                "dm_last": last,
                "dm_role": parsed.get("role", "").strip().replace("_", " ").title(),
                "dm_confidence": parsed.get("confidence", "medium").strip(),
                "dm_source": "Web search + LLM",
            }
        return None

    except (json.JSONDecodeError, anthropic.APIError, KeyError, IndexError) as e:
        logger.debug("LLM parse failed for '%s': %s", entity_name, e)
        return None


def _is_fully_resolved(buyer: dict) -> bool:
    """Check if a buyer has a real decision-maker identified (not surname-only guess)."""
    dm = buyer.get("dm_full", "")
    if not dm or dm in ("Not Found", "Needs Research"):
        return False
    # Surname-only with low confidence = not fully resolved
    conf = buyer.get("dm_confidence", "")
    if conf == "low" and " " not in dm:
        return False
    return True


def _research_single_buyer(buyer: dict, api_key: str, count: int) -> bool:
    """Run all three tiers of resolution on a single buyer. Returns True if resolved."""
    name = buyer.get("BuyerFullName_original") or buyer.get("BuyerFullName", "")
    entity_type = buyer.get("EntityType", "UNKNOWN")

    # Tier 1: Individuals
    if entity_type == "INDIVIDUAL":
        result = _parse_individual_name(name)
        buyer.update(result)
        return True

    # Tier 2: Name parsing (free) — only accept high/medium confidence
    parsed = _try_parse_entity_name(name, entity_type)
    if parsed and parsed.get("dm_confidence") in ("high", "medium"):
        buyer.update(parsed)
        return True

    # Tier 3: Web search + LLM
    if api_key:
        logger.info("  [%d] Researching: %s", count, name)
        time.sleep(random.uniform(SEARCH_DELAY_MIN, SEARCH_DELAY_MAX))

        state = (buyer.get("BuyerState") or "TN").strip()
        if state == "TN":
            state = "Tennessee"

        search_results = _search_entity(name, state)
        if search_results:
            llm_result = _parse_with_llm(name, entity_type, search_results, api_key)
            if llm_result:
                buyer.update(llm_result)
                return True

        # Fallback: try alternate search queries
        alt_queries = [
            f'"{name}" tnbear.tn.gov',
            f'"{name}" Knoxville Tennessee owner',
            f'"{name}" site:bizapedia.com OR site:opencorporates.com',
        ]
        for alt_q in alt_queries:
            time.sleep(random.uniform(SEARCH_DELAY_MIN, SEARCH_DELAY_MAX))
            try:
                alt_results = DDGS().text(alt_q, max_results=5)
                alt_filtered = [
                    {"url": r.get("href", ""), "title": r.get("title", ""),
                     "snippet": r.get("body", "")}
                    for r in alt_results if r.get("href")
                ]
                if alt_filtered:
                    llm_result = _parse_with_llm(name, entity_type, alt_filtered, api_key)
                    if llm_result:
                        buyer.update(llm_result)
                        return True
            except Exception:
                pass

        # Low-confidence name parse as last resort (better than nothing)
        if parsed:
            buyer.update(parsed)
            return False  # Still not "fully" resolved

        buyer["dm_full"] = "Not Found"
        buyer["dm_first"] = ""
        buyer["dm_last"] = ""
        buyer["dm_role"] = ""
        buyer["dm_confidence"] = ""
        buyer["dm_source"] = "Web search (no results)"
    else:
        # Use low-confidence name parse if available
        if parsed:
            buyer.update(parsed)
        else:
            buyer["dm_full"] = "Needs Research"
            buyer["dm_first"] = ""
            buyer["dm_last"] = ""
            buyer["dm_role"] = ""
            buyer["dm_confidence"] = ""
            buyer["dm_source"] = ""

    return False


def resolve_decision_makers(buyers: list[dict], target_resolved: int = 25,
                            use_web_search: bool = False) -> list[dict]:
    """Resolve decision-makers, continuing down the list until target is met.

    Keeps researching beyond the initial batch until we have enough
    fully-resolved buyers to fill the top N list.
    """
    api_key = config.ANTHROPIC_API_KEY if use_web_search else ""
    research_count = 0
    resolved_count = 0
    max_to_research = min(len(buyers), target_resolved * 3)  # Cap at 3x target

    for buyer in buyers[:max_to_research]:
        research_count += 1
        if _research_single_buyer(buyer, api_key, research_count):
            resolved_count += 1

        # Stop once we have enough fully resolved
        if resolved_count >= target_resolved + 5:  # Small buffer
            break

    logger.info("Resolved %d/%d decision-makers (%d buyers processed)",
                resolved_count, target_resolved, research_count)

    # Re-sort: fully resolved buyers first (by score), then unresolved
    resolved = [b for b in buyers if _is_fully_resolved(b)]
    unresolved = [b for b in buyers if not _is_fully_resolved(b)]
    resolved.sort(key=lambda b: b.get("score_total", 0), reverse=True)
    unresolved.sort(key=lambda b: b.get("score_total", 0), reverse=True)

    reordered = resolved + unresolved
    for i, b in enumerate(reordered):
        b["rank"] = i + 1

    return reordered


# ── Excel Output ─────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_RED_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
_LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _write_headers(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_widths(ws, min_w=12, max_w=35):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(mx + 2, min_w), max_w)


def _dm_fill(dm_full: str) -> PatternFill:
    """Color-code decision-maker cell by resolution status."""
    if not dm_full or dm_full in ("Not Found", "Needs Research"):
        return _RED_FILL
    if " " not in dm_full:  # Surname only
        return _YELLOW_FILL
    return _GREEN_FILL


def generate_excel(buyers: list[dict], excluded: list[dict],
                   county: str, state: str, top_n: int,
                   output_path: str = "") -> str:
    """Generate the polished multi-tab Excel workbook."""
    wb = Workbook()
    today = datetime.now().strftime("%Y-%m-%d")

    # ── Tab 1: Top N Cash Buyers ─────────────────────────────────────
    ws = wb.active
    ws.title = f"Top {top_n} Cash Buyers"
    ws.cell(row=1, column=1,
            value=f"Top {top_n} Wholesale-Friendly Cash Buyers").font = _TITLE_FONT
    ws.cell(row=2, column=1,
            value=f"{county} County, {state} | Generated {today}").font = _SUBTITLE_FONT
    ws.cell(row=3, column=1,
            value="Scored for wholesale deal fit. Institutional/government buyers excluded.").font = _LABEL_FONT

    headers = [
        "Rank", "Buyer Name", "Decision Maker", "DM First", "DM Last",
        "Role", "Score", "Purchases (6mo)", "Entity Type",
        "City", "State", "Address", "ZIP", "Confidence", "Source",
    ]
    _write_headers(ws, 5, headers)

    for i, buyer in enumerate(buyers[:top_n], 6):
        name = buyer.get("BuyerFullName_original") or buyer.get("BuyerFullName", "")
        dm_full = buyer.get("dm_full", "")
        ws.cell(row=i, column=1, value=buyer.get("rank", i - 5))
        ws.cell(row=i, column=2, value=name)
        dm_cell = ws.cell(row=i, column=3, value=dm_full)
        dm_cell.fill = _dm_fill(dm_full)
        ws.cell(row=i, column=4, value=buyer.get("dm_first", ""))
        ws.cell(row=i, column=5, value=buyer.get("dm_last", ""))
        ws.cell(row=i, column=6, value=buyer.get("dm_role", ""))
        ws.cell(row=i, column=7, value=buyer.get("score_total", 0))
        ws.cell(row=i, column=8, value=buyer["BuyerPurchases6MSum"])
        ws.cell(row=i, column=9, value=buyer.get("EntityType", ""))
        ws.cell(row=i, column=10, value=buyer.get("BuyerCity", ""))
        ws.cell(row=i, column=11, value=buyer.get("BuyerState", ""))
        ws.cell(row=i, column=12, value=buyer.get("BuyerAddress", ""))
        ws.cell(row=i, column=13, value=buyer.get("BuyerZIP", ""))
        ws.cell(row=i, column=14, value=buyer.get("dm_confidence", ""))
        ws.cell(row=i, column=15, value=buyer.get("dm_source", ""))
        for c in range(1, 16):
            ws.cell(row=i, column=c).border = _THIN_BORDER

    ws.freeze_panes = "A6"
    _auto_widths(ws)

    # ── Tab 2: Scoring Detail ────────────────────────────────────────
    ws2 = wb.create_sheet("Scoring Detail")
    ws2.cell(row=1, column=1, value="Scoring Breakdown").font = _TITLE_FONT
    ws2.cell(row=2, column=1,
             value="Local (30) + Activity (25) + Entity (20) + Name (15) + Address (10) = Total (100)").font = _LABEL_FONT

    score_headers = [
        "Rank", "Buyer", "Local (30)", "Activity (25)", "Entity (20)",
        "Name (15)", "Address (10)", "Total (100)",
    ]
    _write_headers(ws2, 4, score_headers)

    for i, buyer in enumerate(buyers[:top_n], 5):
        name = buyer.get("BuyerFullName_original") or buyer.get("BuyerFullName", "")
        ws2.cell(row=i, column=1, value=buyer.get("rank", i - 4))
        ws2.cell(row=i, column=2, value=name)
        ws2.cell(row=i, column=3, value=buyer.get("score_local", 0))
        ws2.cell(row=i, column=4, value=buyer.get("score_activity", 0))
        ws2.cell(row=i, column=5, value=buyer.get("score_entity", 0))
        ws2.cell(row=i, column=6, value=buyer.get("score_name", 0))
        ws2.cell(row=i, column=7, value=buyer.get("score_address", 0))
        ws2.cell(row=i, column=8, value=buyer.get("score_total", 0))
        for c in range(1, 9):
            ws2.cell(row=i, column=c).border = _THIN_BORDER

    _auto_widths(ws2)

    # ── Tab 3: Excluded Buyers ───────────────────────────────────────
    ws3 = wb.create_sheet("Excluded Buyers")
    ws3.cell(row=1, column=1, value="Excluded Buyers (Audit Trail)").font = _TITLE_FONT
    ws3.cell(row=2, column=1,
             value=f"{len(excluded)} buyers excluded as institutional, government, or too large").font = _LABEL_FONT

    excl_headers = ["Buyer Name", "Purchases (6mo)", "City", "State", "Exclusion Reason"]
    _write_headers(ws3, 4, excl_headers)

    for i, buyer in enumerate(excluded, 5):
        name = buyer.get("BuyerFullName_original") or buyer.get("BuyerFullName", "")
        ws3.cell(row=i, column=1, value=name)
        ws3.cell(row=i, column=2, value=buyer["BuyerPurchases6MSum"])
        ws3.cell(row=i, column=3, value=buyer.get("BuyerCity", ""))
        ws3.cell(row=i, column=4, value=buyer.get("BuyerState", ""))
        ws3.cell(row=i, column=5, value=buyer.get("exclusion_reason", ""))
        for c in range(1, 6):
            ws3.cell(row=i, column=c).border = _THIN_BORDER

    _auto_widths(ws3)

    # ── Tab 4: All Remaining ─────────────────────────────────────────
    ws4 = wb.create_sheet("All Remaining")
    ws4.cell(row=1, column=1,
             value=f"All {len(buyers)} Non-Excluded Buyers (Ranked)").font = _TITLE_FONT

    all_headers = [
        "Rank", "Buyer Name", "Score", "Purchases (6mo)", "Entity Type",
        "City", "State", "ZIP",
    ]
    _write_headers(ws4, 3, all_headers)

    for i, buyer in enumerate(buyers, 4):
        name = buyer.get("BuyerFullName_original") or buyer.get("BuyerFullName", "")
        ws4.cell(row=i, column=1, value=buyer.get("rank", i - 3))
        ws4.cell(row=i, column=2, value=name)
        ws4.cell(row=i, column=3, value=buyer.get("score_total", 0))
        ws4.cell(row=i, column=4, value=buyer["BuyerPurchases6MSum"])
        ws4.cell(row=i, column=5, value=buyer.get("EntityType", ""))
        ws4.cell(row=i, column=6, value=buyer.get("BuyerCity", ""))
        ws4.cell(row=i, column=7, value=buyer.get("BuyerState", ""))
        ws4.cell(row=i, column=8, value=buyer.get("BuyerZIP", ""))
        for c in range(1, 9):
            ws4.cell(row=i, column=c).border = _THIN_BORDER

    _auto_widths(ws4)

    # Save
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d")
        output_path = str(
            config.OUTPUT_DIR
            / f"{county}_County_Top{top_n}_CashBuyers_{timestamp}.xlsx"
        )

    wb.save(output_path)
    logger.info("Saved buyer list to %s", output_path)
    return output_path


# ── Main ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Generate a Top N wholesale-friendly cash buyer list"
    )
    parser.add_argument("--county", required=True, help="County name (e.g. Knox)")
    parser.add_argument("--state", required=True, help="State abbrev (e.g. TN)")
    parser.add_argument("--top", type=int, default=25, help="Number of buyers (default 25)")
    parser.add_argument("--research", action="store_true",
                        help="Enable web search + LLM for decision-maker research")
    parser.add_argument("--output", default="", help="Output Excel path")
    parser.add_argument("--min-purchases", type=int, default=2,
                        help="Minimum purchases to include (default 2)")
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbose logging")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-8s %(message)s",
        datefmt="%H:%M:%S",
    )

    print(f"\n{'='*60}")
    print(f"  Cash Buyer List: {args.county} County, {args.state}")
    print(f"  Top {args.top} | Min purchases: {args.min_purchases}")
    print(f"  Web research: {'ON' if args.research else 'OFF'}")
    print(f"{'='*60}\n")

    # Step 1: Load and filter
    buyers = load_and_filter(args.county, args.state, args.min_purchases)
    if not buyers:
        print("No buyers found. Check county/state.")
        return

    # Step 2: Exclude institutional
    buyers, excluded = exclude_institutional(buyers)
    print(f"After exclusions: {len(buyers)} kept, {len(excluded)} excluded")

    # Step 3: Categorize entities
    for buyer in buyers:
        name = buyer.get("BuyerFullName_original") or buyer.get("BuyerFullName", "")
        buyer["EntityType"] = categorize_entity(name)

    entity_dist = {}
    for b in buyers:
        t = b["EntityType"]
        entity_dist[t] = entity_dist.get(t, 0) + 1
    print(f"\nEntity types: {entity_dist}")

    # Step 4: Score for wholesale-friendliness
    buyers = score_buyers(buyers)
    print(f"\nTop 10 by score:")
    for b in buyers[:10]:
        name = b.get("BuyerFullName_original") or b.get("BuyerFullName", "")
        print(f"  {b['rank']:>3}. {name:<45} Score: {b['score_total']:>5.1f}  "
              f"Purchases: {b['BuyerPurchases6MSum']:>3}  "
              f"({b.get('BuyerCity', '')}, {b.get('BuyerState', '')})")

    # Step 5: Resolve decision-makers — keep going until we have enough
    print(f"\nResolving decision-makers (target: {args.top} fully uncovered)...")
    buyers = resolve_decision_makers(buyers, args.top, args.research)

    resolved = sum(1 for b in buyers[:args.top] if _is_fully_resolved(b))
    print(f"Fully resolved: {resolved}/{args.top} decision-makers in top {args.top}")

    # Step 6: Generate Excel
    output = generate_excel(buyers, excluded, args.county, args.state, args.top, args.output)
    print(f"\nOutput: {output}")

    # Summary
    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"  Total Knox County buyers: {len(buyers) + len(excluded)}")
    print(f"  Excluded (institutional): {len(excluded)}")
    print(f"  Scored & ranked: {len(buyers)}")
    print(f"  Decision-makers resolved: {resolved}/{args.top}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
