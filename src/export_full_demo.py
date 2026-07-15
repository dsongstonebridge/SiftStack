"""Export ALL scraped records from master CSV to a presentation-ready Excel workbook.

Showcases the full pipeline: scrape → parse → enrich → heir research.
Produces 6 sheets highlighting enrichment coverage, best prospects, and deceased owner DM research.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MASTER_CSV = Path("output/master_all_records.csv")
OUT_DIR = Path("output")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SECTION_FONT = Font(bold=True, size=11, color="1F4E78")
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")


def format_owner(name: str | float) -> str:
    if not isinstance(name, str) or not name.strip():
        return ""
    parts = name.strip().split()
    if len(parts) >= 2 and parts[0].isupper() and not any(c.islower() for c in parts[0]):
        return " ".join(parts[1:] + [parts[0]])
    return name.strip()


def style_sheet(
    ws,
    df: pd.DataFrame,
    title: str,
    currency_cols: list[str] | None = None,
    percent_cols: list[str] | None = None,
    url_cols: list[str] | None = None,
    highlight_col: str | None = None,
) -> None:
    currency_cols = currency_cols or []
    percent_cols = percent_cols or []
    url_cols = url_cols or []

    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    if len(df.columns) > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
        is_zebra = row_idx % 2 == 0
        highlight_row = False
        if highlight_col and highlight_col in df.columns:
            val = row[highlight_col]
            highlight_row = isinstance(val, str) and val.lower() == "yes"

        for col_idx, col_name in enumerate(df.columns, start=1):
            value = row[col_name]
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in currency_cols and isinstance(value, (int, float)) and value != "":
                cell.number_format = '"$"#,##0'
            if col_name in percent_cols and isinstance(value, (int, float)) and value != "":
                cell.number_format = '0"%"'
            if col_name in url_cols and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
                cell.value = "View Notice"
            if highlight_row:
                cell.fill = HIGHLIGHT_FILL
            elif is_zebra:
                cell.fill = ZEBRA_FILL

    for col_idx, col_name in enumerate(df.columns, start=1):
        series = df[col_name].astype(object).where(df[col_name].notna(), "").astype(str)
        max_len = max([len(str(col_name))] + [len(s) for s in series])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 45)

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    rows: list[tuple[str, str]] = []

    rows.append(("PIPELINE OVERVIEW", ""))
    rows.append(("Total Records", f"{total:,}"))
    rows.append(("Counties", ", ".join(sorted(df["county"].dropna().unique()))))
    rows.append(("Notice Types", ", ".join(sorted(df["notice_type"].dropna().unique()))))
    rows.append(("Date Range", f"{df['auction_date'].dropna().min()} — {df['auction_date'].dropna().max()}"))
    rows.append(("Source", "tnpublicnotice.com (Playwright + 2Captcha)"))
    rows.append(("", ""))

    rows.append(("BREAKDOWN BY COUNTY × TYPE", ""))
    for (county, ntype), count in df.groupby(["county", "notice_type"]).size().items():
        rows.append((f"  {county} — {ntype}", f"{count:,}"))
    rows.append(("", ""))

    rows.append(("ENRICHMENT COVERAGE", ""))
    coverage_fields = [
        ("Address standardized", "dpv_match_code"),
        ("Property valuation (Zillow)", "estimated_value"),
        ("Equity calculation", "estimated_equity"),
        ("Last sale history", "mls_last_sold_date"),
        ("Property attributes (beds/baths/sqft)", "sqft"),
        ("Vacancy flag", "vacant"),
        ("Auction date", "auction_date"),
        ("Parcel ID (tax records)", "parcel_id"),
        ("Tax delinquent amount", "tax_delinquent_amount"),
        ("Deceased owner identified", "owner_deceased"),
        ("Decision Maker researched", "decision_maker_name"),
        ("Obituary matched", "obituary_url"),
    ]
    for label, col in coverage_fields:
        if col in df.columns:
            if col == "owner_deceased":
                count = (df[col] == "yes").sum()
            else:
                count = df[col].notna().sum()
            pct = (count / total * 100) if total else 0
            rows.append((f"  {label}", f"{count:,} ({pct:.0f}%)"))
    rows.append(("", ""))

    value_df = df[df["estimated_value"].notna()]
    if len(value_df):
        rows.append(("PORTFOLIO VALUE", ""))
        rows.append(("  Total Estimated Value", f"${value_df['estimated_value'].sum():,.0f}"))
        rows.append(("  Average Estimated Value", f"${value_df['estimated_value'].mean():,.0f}"))
        equity_df = df[df["estimated_equity"].notna()]
        if len(equity_df):
            rows.append(("  Total Equity", f"${equity_df['estimated_equity'].sum():,.0f}"))
            high_eq = equity_df[equity_df["equity_percent"] >= 50]
            rows.append(("  High-Equity Deals (≥50%)", f"{len(high_eq):,}"))
        rows.append(("", ""))

    rows.append(("GENERATED", datetime.now().strftime("%Y-%m-%d %H:%M")))

    return pd.DataFrame(rows, columns=["Metric", "Value"])


def clean_records(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["full_name"] = df["full_name"].apply(format_owner)
    if "zip" in df.columns:
        df["zip"] = df["zip"].astype("Int64").astype(str).replace("<NA>", "")
    for col in ["bedrooms", "bathrooms", "year_built", "tax_delinquent_years"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").round().astype("Int64")
    return df


def all_records_view(df: pd.DataFrame) -> pd.DataFrame:
    cols = {
        "notice_type": "Type",
        "county": "County",
        "address": "Property Address",
        "city": "City",
        "zip": "ZIP",
        "full_name": "Owner",
        "auction_date": "Auction Date",
        "estimated_value": "Est. Value",
        "estimated_equity": "Equity $",
        "equity_percent": "Equity %",
        "bedrooms": "Beds",
        "bathrooms": "Baths",
        "sqft": "SqFt",
        "year_built": "Year Built",
        "vacant": "Vacant",
        "owner_deceased": "Deceased",
        "source_url": "Notice",
    }
    available = {k: v for k, v in cols.items() if k in df.columns}
    return df[list(available.keys())].rename(columns=available).sort_values(
        ["County", "Type", "Auction Date"], na_position="last"
    )


def best_prospects_view(df: pd.DataFrame) -> pd.DataFrame:
    prospects = df[
        (df["estimated_equity"].notna()) & (df["estimated_value"].notna())
    ].copy()
    prospects["priority_score"] = (
        prospects["estimated_equity"].fillna(0)
        + (prospects["owner_deceased"] == "yes").astype(int) * 50000
        + (prospects["vacant"] == "Y").astype(int) * 25000
    )
    prospects = prospects.sort_values("priority_score", ascending=False).head(50)

    cols = {
        "notice_type": "Type",
        "county": "County",
        "address": "Property Address",
        "city": "City",
        "full_name": "Owner",
        "estimated_value": "Est. Value",
        "estimated_equity": "Equity $",
        "equity_percent": "Equity %",
        "bedrooms": "Beds",
        "bathrooms": "Baths",
        "sqft": "SqFt",
        "vacant": "Vacant",
        "owner_deceased": "Deceased",
        "decision_maker_name": "Decision Maker",
        "auction_date": "Auction Date",
        "source_url": "Notice",
    }
    available = {k: v for k, v in cols.items() if k in prospects.columns}
    return prospects[list(available.keys())].rename(columns=available)


def tax_sale_view(df: pd.DataFrame) -> pd.DataFrame:
    tax = df[df["notice_type"] == "tax_sale"].copy()
    cols = {
        "address": "Property Address",
        "city": "City",
        "zip": "ZIP",
        "full_name": "Owner",
        "parcel_id": "Parcel ID",
        "tax_delinquent_amount": "Amount Owed",
        "tax_delinquent_years": "Years Delinquent",
        "estimated_value": "Est. Value",
        "owner_deceased": "Deceased",
        "decision_maker_name": "Decision Maker",
        "source_url": "Notice",
    }
    available = {k: v for k, v in cols.items() if k in tax.columns}
    return tax[list(available.keys())].rename(columns=available).sort_values(
        "Amount Owed", ascending=False, na_position="last"
    )


def deceased_view(df: pd.DataFrame) -> pd.DataFrame:
    deceased = df[df["owner_deceased"] == "yes"].copy()
    cols = {
        "notice_type": "Type",
        "county": "County",
        "address": "Property Address",
        "city": "City",
        "full_name": "Deceased Owner",
        "date_of_death": "Date of Death",
        "decision_maker_name": "Decision Maker (DM)",
        "decision_maker_relationship": "DM Relationship",
        "decision_maker_street": "DM Street",
        "decision_maker_city": "DM City",
        "decision_maker_state": "DM State",
        "decision_maker_zip": "DM ZIP",
        "dm_confidence": "Confidence",
        "obituary_url": "Obituary",
        "source_url": "Notice",
    }
    available = {k: v for k, v in cols.items() if k in deceased.columns}
    return deceased[list(available.keys())].rename(columns=available)


def main() -> Path:
    raw = pd.read_csv(MASTER_CSV, low_memory=False)
    df = clean_records(raw)

    summary_df = build_summary(df)
    all_df = all_records_view(df)
    prospects_df = best_prospects_view(df)
    tax_df = tax_sale_view(df)
    deceased_df = deceased_view(df)

    out_path = OUT_DIR / f"full_pipeline_demo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=2, header=False)
        all_df.to_excel(writer, sheet_name="All Records", index=False, startrow=2, header=False)
        prospects_df.to_excel(writer, sheet_name="Best Prospects", index=False, startrow=2, header=False)
        tax_df.to_excel(writer, sheet_name="Tax Sale", index=False, startrow=2, header=False)
        deceased_df.to_excel(writer, sheet_name="Deceased + DM Research", index=False, startrow=2, header=False)

        wb = writer.book
        style_sheet(wb["Summary"], summary_df, "TN Public Notice Pipeline — Summary")
        style_sheet(
            wb["All Records"],
            all_df,
            f"All Records — {len(all_df):,} Properties",
            currency_cols=["Est. Value", "Equity $"],
            percent_cols=["Equity %"],
            url_cols=["Notice"],
            highlight_col="Deceased",
        )
        style_sheet(
            wb["Best Prospects"],
            prospects_df,
            "Top 50 Prospects — Ranked by Equity + Deceased + Vacancy",
            currency_cols=["Est. Value", "Equity $"],
            percent_cols=["Equity %"],
            url_cols=["Notice"],
            highlight_col="Deceased",
        )
        style_sheet(
            wb["Tax Sale"],
            tax_df,
            f"Knox County Tax Sale — {len(tax_df)} Properties",
            currency_cols=["Amount Owed", "Est. Value"],
            url_cols=["Notice"],
            highlight_col="Deceased",
        )
        style_sheet(
            wb["Deceased + DM Research"],
            deceased_df,
            f"Deceased Owners with Heir/DM Research — {len(deceased_df)} Records",
            url_cols=["Notice", "Obituary"],
        )

    print(f"Wrote {out_path}")
    print(f"  All records: {len(all_df):,}")
    print(f"  Best prospects: {len(prospects_df):,}")
    print(f"  Tax sale: {len(tax_df):,}")
    print(f"  Deceased + DM: {len(deceased_df):,}")
    return out_path


if __name__ == "__main__":
    main()
