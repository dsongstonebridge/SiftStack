"""Export tax sale records from master CSV to a cleaned, presentation-ready Excel file.

Takes the raw 80+ column master CSV and produces a focused 3-sheet workbook
highlighting the tax sale list: Summary, Property List, Decision Makers.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MASTER_CSV = Path("output/master_all_records.csv")
OUT_DIR = Path("output")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")


def format_owner(name: str | float) -> str:
    if not isinstance(name, str) or not name.strip():
        return ""
    parts = name.strip().split()
    if len(parts) >= 2 and parts[0].isupper() and not any(c.islower() for c in parts[0]):
        return " ".join(parts[1:] + [parts[0]])
    return name.strip()


def style_sheet(ws, df: pd.DataFrame, title: str, currency_cols: list[str] = None, url_cols: list[str] = None) -> None:
    currency_cols = currency_cols or []
    url_cols = url_cols or []

    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
        is_zebra = row_idx % 2 == 0
        for col_idx, col_name in enumerate(df.columns, start=1):
            value = row[col_name]
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in currency_cols and isinstance(value, (int, float)) and value != "":
                cell.number_format = '"$"#,##0.00'
            if col_name in url_cols and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
                cell.value = "View Notice"
            if is_zebra:
                cell.fill = ZEBRA_FILL

    for col_idx, col_name in enumerate(df.columns, start=1):
        series = df[col_name].fillna("").astype(str)
        max_len = max([len(col_name)] + [len(s) for s in series])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 45)

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


def build_summary(tax: pd.DataFrame) -> pd.DataFrame:
    total_owed = tax["tax_delinquent_amount"].sum()
    avg_owed = tax["tax_delinquent_amount"].mean()
    max_years = tax["tax_delinquent_years"].max()
    deceased_count = (tax["owner_deceased"] == "yes").sum()
    dm_identified = tax["decision_maker_name"].notna().sum()

    rows = [
        ("Total Records", f"{len(tax)}"),
        ("Counties", ", ".join(sorted(tax["county"].dropna().unique()))),
        ("ZIP Codes", ", ".join(sorted(tax["zip"].dropna().astype(int).astype(str).unique()))),
        ("Total Delinquent", f"${total_owed:,.2f}"),
        ("Average Delinquent", f"${avg_owed:,.2f}"),
        ("Longest Delinquency (years)", f"{int(max_years) if pd.notna(max_years) else 0}"),
        ("Deceased Owners Identified", f"{deceased_count}"),
        ("Decision Makers Identified", f"{dm_identified}"),
        ("Source", "tnpublicnotice.com (scraper.py)"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def main() -> Path:
    df = pd.read_csv(MASTER_CSV, low_memory=False)
    tax = df[df["notice_type"] == "tax_sale"].copy()
    tax["full_name"] = tax["full_name"].apply(format_owner)
    tax["zip"] = tax["zip"].astype("Int64").astype(str)

    property_cols = {
        "address": "Property Address",
        "city": "City",
        "zip": "ZIP",
        "full_name": "Owner",
        "parcel_id": "Parcel ID",
        "tax_delinquent_amount": "Amount Owed",
        "tax_delinquent_years": "Years Delinquent",
        "owner_deceased": "Deceased",
        "source_url": "Notice",
    }
    property_df = tax[list(property_cols.keys())].rename(columns=property_cols)
    property_df = property_df.sort_values("Amount Owed", ascending=False, na_position="last")

    dm_cols = {
        "address": "Property Address",
        "full_name": "Owner",
        "owner_deceased": "Deceased",
        "date_of_death": "Date of Death",
        "decision_maker_name": "Decision Maker",
        "decision_maker_relationship": "Relationship",
        "decision_maker_street": "DM Street",
        "decision_maker_city": "DM City",
        "decision_maker_state": "DM State",
        "decision_maker_zip": "DM ZIP",
        "dm_confidence": "Confidence",
    }
    dm_df = tax[tax["decision_maker_name"].notna()][list(dm_cols.keys())].rename(columns=dm_cols)

    summary_df = build_summary(tax)

    out_path = OUT_DIR / f"tax_sale_demo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=2, header=False)
        property_df.to_excel(writer, sheet_name="Tax Sale List", index=False, startrow=2, header=False)
        dm_df.to_excel(writer, sheet_name="Decision Makers", index=False, startrow=2, header=False)

        wb = writer.book
        style_sheet(wb["Summary"], summary_df, "Knox County Tax Sale — Summary")
        style_sheet(
            wb["Tax Sale List"],
            property_df,
            "Knox County Tax Sale — Property List",
            currency_cols=["Amount Owed"],
            url_cols=["Notice"],
        )
        style_sheet(
            wb["Decision Makers"],
            dm_df,
            "Knox County Tax Sale — Decision Makers (Deceased Owners)",
        )

    print(f"Wrote {out_path} ({len(tax)} records)")
    return out_path


if __name__ == "__main__":
    main()
