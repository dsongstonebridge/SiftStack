"""Lead management system with 4 Pillars of Motivation scoring and STABM routine.

Qualifies leads automatically from enrichment data, routes by temperature,
and generates daily STABM (Status-Task-Board-Message) reports.

4 Pillars of Motivation:
  1. Reason — why would they sell to an investor? (distress level)
  2. Timeline — how urgently do they need to sell?
  3. Condition — what state is the property in?
  4. Price — how does asking compare to Zestimate?

2+ Hot pillars → Closer queue. 2+ Cold → Long-term drip.

Usage:
  python src/main.py lead-manage --action qualify --csv-path output/records.csv
  python src/main.py lead-manage --action report
"""

import csv
import logging
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config

logger = logging.getLogger(__name__)

# ── Temperature thresholds ────────────────────────────────────────────
# Hot: clear distress, <30 days, major work, ≤80% Zestimate
# Warm: moderate interest, 1-3 months, some updates, 80-100%
# Cold: testing market, 3-12+ months, good condition, >100%

AUCTION_HOT_DAYS = 30
AUCTION_WARM_DAYS = 90
YEAR_BUILT_HOT = 1970    # Built before = major work likely
YEAR_BUILT_WARM = 1990
PRICE_HOT_PCT = 80       # ≤80% of Zestimate
PRICE_WARM_PCT = 100


@dataclass
class PillarScore:
    """Score for a single pillar."""
    name: str = ""
    temperature: str = ""  # "hot", "warm", "cold"
    reason: str = ""
    score: int = 0  # 3=hot, 2=warm, 1=cold


@dataclass
class LeadQualification:
    """Full qualification for a single lead."""
    address: str = ""
    owner_name: str = ""
    notice_type: str = ""
    county: str = ""
    reason_pillar: PillarScore = field(default_factory=lambda: PillarScore(name="Reason"))
    timeline_pillar: PillarScore = field(default_factory=lambda: PillarScore(name="Timeline"))
    condition_pillar: PillarScore = field(default_factory=lambda: PillarScore(name="Condition"))
    price_pillar: PillarScore = field(default_factory=lambda: PillarScore(name="Price"))
    overall_temperature: str = ""  # "hot", "warm", "cold"
    hot_count: int = 0
    route_to: str = ""  # "closer", "nurture", "drip", "deep_prospecting"
    score_total: int = 0  # 4-12 scale


@dataclass
class PipelineReport:
    """Daily STABM pipeline report."""
    date: str = ""
    total_leads: int = 0
    hot_leads: int = 0
    warm_leads: int = 0
    cold_leads: int = 0
    by_type: dict = field(default_factory=dict)
    by_route: dict = field(default_factory=dict)
    by_county: dict = field(default_factory=dict)
    qualifications: list = field(default_factory=list)
    stale_records: list = field(default_factory=list)  # >7 days without update
    pipeline_funnel: dict = field(default_factory=dict)


# ── Qualification engine ──────────────────────────────────────────────

def _score_reason(row: dict) -> PillarScore:
    """Score the Reason pillar — why would they sell to an investor?"""
    notice_type = (row.get("notice_type") or row.get("Notice Type") or "").lower()
    deceased = (row.get("owner_deceased") or "").lower() == "yes"
    tax_del = row.get("tax_delinquent_amount") or row.get("Tax Delinquent Value") or ""
    has_tax = False
    if tax_del:
        try:
            has_tax = float(str(tax_del).replace(",", "").replace("$", "")) > 0
        except ValueError:
            pass

    # Hot: foreclosure, tax sale, probate (deceased), high tax delinquency
    hot_types = {"foreclosure", "tax_sale", "probate", "code_violation"}
    if notice_type in hot_types or deceased or has_tax:
        reason = []
        if notice_type in hot_types:
            reason.append(f"distress type: {notice_type}")
        if deceased:
            reason.append("owner deceased")
        if has_tax:
            reason.append(f"tax delinquent: ${tax_del}")
        return PillarScore(name="Reason", temperature="hot", score=3,
                           reason="; ".join(reason))

    # Warm: tax delinquent (no sale), eviction, divorce
    warm_types = {"tax_delinquent", "eviction", "divorce"}
    if notice_type in warm_types:
        return PillarScore(name="Reason", temperature="warm", score=2,
                           reason=f"moderate distress: {notice_type}")

    return PillarScore(name="Reason", temperature="cold", score=1,
                       reason="no clear distress signal")


def _norm_date(raw: str) -> str:
    """Normalize a date string (ISO or Sift M/D/YYYY) to YYYY-MM-DD, else ''."""
    raw = (raw or "").strip()
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(raw[:10] if fmt == "%Y-%m-%d" else raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def _score_timeline(row: dict) -> PillarScore:
    """Score the Timeline pillar — how urgently do they need to sell?"""
    auction_date = _norm_date(row.get("auction_date") or row.get("Foreclosure Date") or "")
    # Freshness = when the notice was PUBLISHED, not when we added the record.
    # date_added is now the run/import date (same for every record in a run), so
    # prefer date_published / "Notice Publish Date", falling back to date_added.
    freshness = _norm_date(
        row.get("date_published") or row.get("Notice Publish Date")
        or row.get("date_added") or row.get("Date Added") or ""
    )

    if auction_date:
        try:
            auction = datetime.strptime(auction_date, "%Y-%m-%d")
            days_until = (auction - datetime.now()).days
            if days_until <= AUCTION_HOT_DAYS:
                return PillarScore(name="Timeline", temperature="hot", score=3,
                                   reason=f"auction in {days_until} days")
            elif days_until <= AUCTION_WARM_DAYS:
                return PillarScore(name="Timeline", temperature="warm", score=2,
                                   reason=f"auction in {days_until} days")
        except ValueError:
            pass

    if freshness:
        try:
            added = datetime.strptime(freshness, "%Y-%m-%d")
            days_since = (datetime.now() - added).days
            if days_since < 14:
                return PillarScore(name="Timeline", temperature="hot", score=3,
                                   reason=f"fresh notice ({days_since} days ago)")
            elif days_since < 60:
                return PillarScore(name="Timeline", temperature="warm", score=2,
                                   reason=f"recent notice ({days_since} days ago)")
        except ValueError:
            pass

    return PillarScore(name="Timeline", temperature="cold", score=1,
                       reason="no urgency indicators")


def _score_condition(row: dict) -> PillarScore:
    """Score the Condition pillar — what state is the property in?"""
    year_built = row.get("year_built") or row.get("Year Built") or ""
    est_val = row.get("estimated_value") or row.get("Estimated Value") or ""
    sqft = row.get("sqft") or row.get("Living SqFt") or ""
    vacant = (row.get("vacant") or "").upper() == "Y"

    year = 0
    if year_built:
        try:
            year = int(year_built)
        except ValueError:
            pass

    # Value per sqft (low = distressed)
    val_per_sqft = 0
    if est_val and sqft:
        try:
            v = float(str(est_val).replace(",", "").replace("$", ""))
            s = float(str(sqft).replace(",", ""))
            if s > 0:
                val_per_sqft = v / s
        except ValueError:
            pass

    if vacant:
        return PillarScore(name="Condition", temperature="hot", score=3,
                           reason="property is vacant")
    if year and year < YEAR_BUILT_HOT:
        return PillarScore(name="Condition", temperature="hot", score=3,
                           reason=f"built {year} — likely needs major work")
    if val_per_sqft and val_per_sqft < 80:
        return PillarScore(name="Condition", temperature="hot", score=3,
                           reason=f"low $/sqft (${val_per_sqft:.0f}) — distressed value")
    if year and year < YEAR_BUILT_WARM:
        return PillarScore(name="Condition", temperature="warm", score=2,
                           reason=f"built {year} — may need updates")

    return PillarScore(name="Condition", temperature="cold", score=1,
                       reason="condition appears acceptable")


def _score_price(row: dict) -> PillarScore:
    """Score the Price pillar — asking vs Zestimate."""
    est_val = row.get("estimated_value") or row.get("Estimated Value") or ""
    equity_pct = row.get("equity_percent") or row.get("Equity Percentage") or ""
    last_sold = row.get("mls_last_sold_price") or row.get("Last Sale Price") or ""

    zestimate = 0
    if est_val:
        try:
            zestimate = float(str(est_val).replace(",", "").replace("$", ""))
        except ValueError:
            pass

    equity = 0
    if equity_pct:
        try:
            equity = float(str(equity_pct).replace("%", ""))
        except ValueError:
            pass

    # High equity = more room for investor discount
    if equity >= 50:
        return PillarScore(name="Price", temperature="hot", score=3,
                           reason=f"{equity:.0f}% equity — room for deep discount")
    if equity >= 25:
        return PillarScore(name="Price", temperature="warm", score=2,
                           reason=f"{equity:.0f}% equity — moderate discount possible")
    if equity > 0:
        return PillarScore(name="Price", temperature="cold", score=1,
                           reason=f"{equity:.0f}% equity — limited discount room")

    return PillarScore(name="Price", temperature="cold", score=1,
                       reason="insufficient pricing data")


def qualify_lead(row: dict) -> LeadQualification:
    """Qualify a single lead across all 4 pillars."""
    qual = LeadQualification(
        address=row.get("address") or row.get("Property Street") or "",
        owner_name=row.get("owner_name") or row.get("full_name") or row.get("Owner Name") or "",
        notice_type=row.get("notice_type") or row.get("Notice Type") or "",
        county=row.get("county") or row.get("County") or "",
    )

    qual.reason_pillar = _score_reason(row)
    qual.timeline_pillar = _score_timeline(row)
    qual.condition_pillar = _score_condition(row)
    qual.price_pillar = _score_price(row)

    pillars = [qual.reason_pillar, qual.timeline_pillar,
               qual.condition_pillar, qual.price_pillar]
    qual.hot_count = sum(1 for p in pillars if p.temperature == "hot")
    qual.score_total = sum(p.score for p in pillars)

    # Overall temperature and routing
    if qual.hot_count >= 2:
        qual.overall_temperature = "hot"
        qual.route_to = "closer"
    elif qual.hot_count >= 1 or qual.score_total >= 8:
        qual.overall_temperature = "warm"
        qual.route_to = "nurture"
    else:
        qual.overall_temperature = "cold"
        qual.route_to = "drip"

    # Special routing for deceased owners
    deceased = (row.get("owner_deceased") or "").lower() == "yes"
    dm = row.get("decision_maker_name") or ""
    if deceased and not dm:
        qual.route_to = "deep_prospecting"

    return qual


# ── Batch qualification ───────────────────────────────────────────────

def qualify_batch(csv_path: str, max_records: int = 0) -> list[LeadQualification]:
    """Qualify all leads in a CSV file."""
    qualifications = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            qual = qualify_lead(row)
            qualifications.append(qual)
            if max_records and len(qualifications) >= max_records:
                break
    return qualifications


# ── STABM report ──────────────────────────────────────────────────────

def generate_stabm_report(qualifications: list[LeadQualification],
                          output_path: str = "") -> str:
    """Generate daily STABM pipeline report as Excel workbook."""
    wb = Workbook()

    # Styles
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=16, color="2F5496")
    subtitle_font = Font(name="Calibri", bold=True, size=12, color="333333")
    label_font = Font(name="Calibri", size=11, color="555555")
    value_font = Font(name="Calibri", bold=True, size=13, color="222222")
    thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))
    temp_colors = {
        "hot": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "warm": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "cold": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    }

    # ── Tab 1: STABM Dashboard ────────────────────────────────────
    ws = wb.active
    ws.title = "STABM Dashboard"
    ws.cell(row=1, column=1, value="Daily STABM Report").font = title_font
    ws.cell(row=2, column=1, value=datetime.now().strftime("%A, %B %d, %Y")).font = subtitle_font

    hot = sum(1 for q in qualifications if q.overall_temperature == "hot")
    warm = sum(1 for q in qualifications if q.overall_temperature == "warm")
    cold = sum(1 for q in qualifications if q.overall_temperature == "cold")

    ws.cell(row=4, column=1, value="S — STATUS").font = subtitle_font
    status_data = [
        ("Total Leads", len(qualifications)),
        ("Hot (→ Closer)", hot),
        ("Warm (→ Nurture)", warm),
        ("Cold (→ Drip)", cold),
    ]
    for i, (label, value) in enumerate(status_data, 5):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=2, value=value).font = value_font

    ws.cell(row=10, column=1, value="Ta — TASKS").font = subtitle_font
    ws.cell(row=11, column=1, value=f"Hot leads needing immediate contact: {hot}").font = label_font
    dp = sum(1 for q in qualifications if q.route_to == "deep_prospecting")
    ws.cell(row=12, column=1, value=f"Records needing deep prospecting: {dp}").font = label_font

    ws.cell(row=14, column=1, value="B — BOARD").font = subtitle_font
    by_route = Counter(q.route_to for q in qualifications)
    row = 15
    for route, count in by_route.most_common():
        ws.cell(row=row, column=1, value=route.replace("_", " ").title()).font = label_font
        ws.cell(row=row, column=2, value=count).font = value_font
        row += 1

    ws.cell(row=row + 1, column=1, value="M — MESSAGES").font = subtitle_font
    ws.cell(row=row + 2, column=1, value="Check DataSift for pending follow-ups and responses").font = label_font

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15

    # ── Tab 2: Hot Leads ──────────────────────────────────────────
    ws2 = wb.create_sheet("Hot Leads")
    ws2.cell(row=1, column=1, value="Hot Leads — Immediate Action Required").font = title_font
    headers = ["Address", "Owner", "Type", "County", "Score", "Hot Pillars",
               "Reason", "Timeline", "Condition", "Price", "Route"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill

    hot_leads = [q for q in qualifications if q.overall_temperature == "hot"]
    hot_leads.sort(key=lambda q: q.score_total, reverse=True)
    for i, q in enumerate(hot_leads, 4):
        vals = [q.address, q.owner_name, q.notice_type, q.county, q.score_total,
                q.hot_count, q.reason_pillar.temperature, q.timeline_pillar.temperature,
                q.condition_pillar.temperature, q.price_pillar.temperature, q.route_to]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            if col >= 7 and col <= 10:
                cell.fill = temp_colors.get(str(val), PatternFill())
            cell.border = thin_border

    # ── Tab 3: All Leads ──────────────────────────────────────────
    ws3 = wb.create_sheet("All Leads")
    ws3.cell(row=1, column=1, value="All Leads — Qualified").font = title_font
    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill

    sorted_quals = sorted(qualifications, key=lambda q: q.score_total, reverse=True)
    for i, q in enumerate(sorted_quals, 4):
        vals = [q.address, q.owner_name, q.notice_type, q.county, q.score_total,
                q.hot_count, q.reason_pillar.temperature, q.timeline_pillar.temperature,
                q.condition_pillar.temperature, q.price_pillar.temperature, q.route_to]
        for col, val in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=col, value=val)
            if col >= 7 and col <= 10:
                cell.fill = temp_colors.get(str(val), PatternFill())
            cell.border = thin_border

    # ── Tab 4: By Type ────────────────────────────────────────────
    ws4 = wb.create_sheet("By Notice Type")
    ws4.cell(row=1, column=1, value="Lead Temperature by Notice Type").font = title_font
    type_counter = defaultdict(lambda: {"hot": 0, "warm": 0, "cold": 0, "total": 0})
    for q in qualifications:
        tc = type_counter[q.notice_type]
        tc[q.overall_temperature] += 1
        tc["total"] += 1

    for col, h in enumerate(["Notice Type", "Hot", "Warm", "Cold", "Total"], 1):
        cell = ws4.cell(row=3, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
    for i, (ntype, counts) in enumerate(sorted(type_counter.items(), key=lambda x: x[1]["total"], reverse=True), 4):
        ws4.cell(row=i, column=1, value=ntype)
        ws4.cell(row=i, column=2, value=counts["hot"])
        ws4.cell(row=i, column=3, value=counts["warm"])
        ws4.cell(row=i, column=4, value=counts["cold"])
        ws4.cell(row=i, column=5, value=counts["total"])

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(config.OUTPUT_DIR / f"stabm_report_{timestamp}.xlsx")

    wb.save(output_path)
    logger.info("STABM report saved to %s", output_path)
    return output_path


# ── Main entry point ──────────────────────────────────────────────────

def run_lead_management(action: str = "qualify", csv_path: str = "",
                        output_path: str = "") -> dict:
    """Run lead management action.

    Actions:
      qualify — score all leads in CSV with 4 Pillars
      report — generate STABM pipeline report
    """
    if action in ("qualify", "report"):
        if not csv_path:
            # Find most recent CSV in output dir
            csvs = sorted(config.OUTPUT_DIR.glob("*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
            if csvs:
                csv_path = str(csvs[0])
                logger.info("Using most recent CSV: %s", csv_path)
            else:
                return {"error": "No CSV path provided and no CSVs in output/"}

        logger.info("Qualifying %d leads from %s", 0, csv_path)
        qualifications = qualify_batch(csv_path)

        if not qualifications:
            return {"error": "No leads to qualify"}

        report_path = generate_stabm_report(qualifications, output_path)

        hot = sum(1 for q in qualifications if q.overall_temperature == "hot")
        warm = sum(1 for q in qualifications if q.overall_temperature == "warm")
        cold = sum(1 for q in qualifications if q.overall_temperature == "cold")

        logger.info("Lead qualification complete: %d total — %d hot, %d warm, %d cold",
                    len(qualifications), hot, warm, cold)

        return {
            "qualifications": qualifications,
            "total": len(qualifications),
            "hot": hot,
            "warm": warm,
            "cold": cold,
            "report_path": report_path,
        }

    return {"error": f"Unknown action: {action}"}
