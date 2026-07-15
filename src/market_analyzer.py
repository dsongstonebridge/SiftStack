"""Market analysis, zip code scoring, and Market Finder reports.

Analyzes county-level data to identify target zip codes for investment.
Scores each zip by distress density, property values, equity, and competition.

Data sources:
  - Our own scraped notice data (aggregated by zip)
  - Zillow API (property values via OpenWeb Ninja)
  - Knox County Tax API (delinquency density)

Usage:
  python src/main.py market-analysis --counties Knox,Blount
  python src/main.py market-analysis --counties Knox --zip-codes 37918,37919,37920
"""

import csv
import glob
import logging
import os
import random
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config

logger = logging.getLogger(__name__)

# ── County zip codes ─────────────────────────────────────────────────
KNOX_ZIPS = [
    "37901", "37902", "37909", "37912", "37914", "37915", "37916", "37917",
    "37918", "37919", "37920", "37921", "37922", "37923", "37924", "37931",
    "37932", "37934", "37938",
]
BLOUNT_ZIPS = [
    "37801", "37803", "37804", "37853", "37882", "37886",
]
TULSA_ZIPS = [
    "74003", "74006", "74008", "74011", "74012", "74014", "74015", "74021",
    "74033", "74037", "74055", "74063", "74066", "74070", "74073",
    "74103", "74104", "74105", "74106", "74107", "74108", "74110", "74112",
    "74114", "74115", "74116", "74119", "74120", "74126", "74127", "74128",
    "74129", "74130", "74131", "74132", "74133", "74134", "74135", "74136",
    "74137", "74145", "74146",
]
COUNTY_ZIPS = {
    "knox": KNOX_ZIPS,
    "blount": BLOUNT_ZIPS,
    "tulsa": TULSA_ZIPS,
}

# ── Scoring weights ───────────────────────────────────────────────────
WEIGHT_DISTRESS_DENSITY = 0.30   # Foreclosure/tax sale/probate count per zip
WEIGHT_MEDIAN_VALUE = 0.20       # Lower median = better for investors
WEIGHT_EQUITY_AVG = 0.15         # Higher equity = more room for deals
WEIGHT_TAX_DELINQUENCY = 0.15   # More delinquency = more distress
WEIGHT_COMPETITION = 0.10        # Lower investor activity = less competition
WEIGHT_DOM_AVG = 0.10            # Higher DOM = more negotiating leverage

# ── Data structures ───────────────────────────────────────────────────


@dataclass
class ZipProfile:
    """Profile for a single zip code."""
    zip_code: str = ""
    county: str = ""
    # Notice/distress data from our scraped records
    total_notices: int = 0
    foreclosure_count: int = 0
    tax_sale_count: int = 0
    tax_delinquent_count: int = 0
    probate_count: int = 0
    eviction_count: int = 0
    code_violation_count: int = 0
    # Property data (aggregated from Zillow)
    median_value: float = 0.0
    avg_equity_pct: float = 0.0
    avg_dom: float = 0.0
    property_count: int = 0
    # Tax delinquency data
    avg_tax_delinquent_amount: float = 0.0
    tax_delinquent_property_count: int = 0
    # Competition (investor activity)
    investor_purchase_count: int = 0
    competition_ratio: float = 0.0  # investor purchases / total sales
    # Calculated score
    score: float = 0.0
    rank: int = 0
    grade: str = ""  # A, B, C, D


@dataclass
class MarketReport:
    """Complete market analysis report."""
    county: str = ""
    analysis_date: str = ""
    total_zips: int = 0
    zip_profiles: list = field(default_factory=list)
    top_zips: list = field(default_factory=list)
    total_notices: int = 0
    avg_median_value: float = 0.0
    avg_equity: float = 0.0
    total_distress: int = 0


# ── Data aggregation from our CSVs ───────────────────────────────────

def _load_notice_data(counties: list[str] | None = None) -> dict[str, ZipProfile]:
    """Aggregate notice data from our output CSVs by zip code."""
    profiles = defaultdict(lambda: ZipProfile())

    csv_files = list(config.OUTPUT_DIR.glob("*.csv"))
    if not csv_files:
        logger.warning("No CSV files found in %s", config.OUTPUT_DIR)
        return dict(profiles)

    county_filter = {c.lower() for c in counties} if counties else None

    for csv_path in csv_files:
        try:
            with open(csv_path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    zip_code = (row.get("zip") or row.get("ZIP") or "").strip()[:5]
                    if not zip_code or not zip_code.isdigit():
                        continue

                    county = (row.get("county") or row.get("County") or "").strip().lower()
                    if county_filter and county not in county_filter:
                        continue

                    notice_type = (row.get("notice_type") or row.get("Notice Type") or "").strip().lower()

                    profile = profiles[zip_code]
                    profile.zip_code = zip_code
                    profile.county = county.title()
                    profile.total_notices += 1

                    type_map = {
                        "foreclosure": "foreclosure_count",
                        "tax_sale": "tax_sale_count",
                        "tax_delinquent": "tax_delinquent_count",
                        "probate": "probate_count",
                        "eviction": "eviction_count",
                        "code_violation": "code_violation_count",
                    }
                    attr = type_map.get(notice_type)
                    if attr:
                        setattr(profile, attr, getattr(profile, attr) + 1)

                    # Aggregate property data
                    est_val = row.get("estimated_value") or row.get("Estimated Value") or ""
                    if est_val:
                        try:
                            val = float(est_val.replace(",", "").replace("$", ""))
                            if val > 0:
                                # Running average
                                old_count = profile.property_count
                                profile.property_count += 1
                                profile.median_value = (
                                    (profile.median_value * old_count + val) / profile.property_count
                                )
                        except ValueError:
                            pass

                    equity_pct = row.get("equity_percent") or row.get("Equity Percentage") or ""
                    if equity_pct:
                        try:
                            pct = float(equity_pct.replace("%", ""))
                            if profile.avg_equity_pct:
                                profile.avg_equity_pct = (profile.avg_equity_pct + pct) / 2
                            else:
                                profile.avg_equity_pct = pct
                        except ValueError:
                            pass

                    tax_amt = row.get("tax_delinquent_amount") or ""
                    if tax_amt:
                        try:
                            amt = float(tax_amt.replace(",", "").replace("$", ""))
                            if amt > 0:
                                profile.tax_delinquent_property_count += 1
                                old = profile.avg_tax_delinquent_amount
                                cnt = profile.tax_delinquent_property_count
                                profile.avg_tax_delinquent_amount = (
                                    (old * (cnt - 1) + amt) / cnt
                                )
                        except ValueError:
                            pass

        except Exception as e:
            logger.debug("Error reading %s: %s", csv_path, e)

    return dict(profiles)


# ── Scoring engine ────────────────────────────────────────────────────

def _normalize(values: list[float], higher_is_better: bool = True) -> list[float]:
    """Normalize values to 0-100 scale."""
    if not values:
        return []
    mn, mx = min(values), max(values)
    if mn == mx:
        return [50.0] * len(values)
    normed = [(v - mn) / (mx - mn) * 100 for v in values]
    if not higher_is_better:
        normed = [100 - n for n in normed]
    return normed


def score_zip_codes(profiles: dict[str, ZipProfile]) -> list[ZipProfile]:
    """Score and rank zip codes for investment potential."""
    if not profiles:
        return []

    zips = list(profiles.values())

    # Normalize each dimension
    distress = _normalize([z.total_notices for z in zips], higher_is_better=True)
    values = _normalize([z.median_value for z in zips], higher_is_better=False)  # Lower = better
    equity = _normalize([z.avg_equity_pct for z in zips], higher_is_better=True)
    tax_del = _normalize([z.avg_tax_delinquent_amount for z in zips], higher_is_better=True)
    competition = _normalize([z.competition_ratio for z in zips], higher_is_better=False)  # Lower = better
    dom = _normalize([z.avg_dom for z in zips], higher_is_better=True)  # Higher DOM = more leverage

    for i, z in enumerate(zips):
        z.score = (
            distress[i] * WEIGHT_DISTRESS_DENSITY +
            values[i] * WEIGHT_MEDIAN_VALUE +
            equity[i] * WEIGHT_EQUITY_AVG +
            tax_del[i] * WEIGHT_TAX_DELINQUENCY +
            competition[i] * WEIGHT_COMPETITION +
            dom[i] * WEIGHT_DOM_AVG
        )

    # Sort by score descending
    zips.sort(key=lambda z: z.score, reverse=True)

    # Assign ranks and grades
    for i, z in enumerate(zips):
        z.rank = i + 1
        pct = z.score
        if pct >= 75:
            z.grade = "A"
        elif pct >= 55:
            z.grade = "B"
        elif pct >= 35:
            z.grade = "C"
        else:
            z.grade = "D"

    return zips


# ── Budget allocation ─────────────────────────────────────────────────

def _allocate_budget(zips: list[ZipProfile], monthly_budget: float = 5000.0,
                     max_zips: int = 5) -> list[tuple[str, float, str]]:
    """Allocate marketing budget across top zip codes by score weight."""
    top = zips[:max_zips]
    if not top:
        return []

    total_score = sum(z.score for z in top)
    if total_score == 0:
        equal = monthly_budget / len(top)
        return [(z.zip_code, round(equal), z.grade) for z in top]

    return [
        (z.zip_code, round(monthly_budget * z.score / total_score), z.grade)
        for z in top
    ]


# ── Excel report ──────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_VALUE_FONT = Font(name="Calibri", bold=True, size=13, color="222222")
_GRADE_COLORS = {
    "A": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "B": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "C": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "D": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
_MONEY_FMT = '#,##0'


def _write_headers(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_widths(ws, min_w=12, max_w=30):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(mx + 2, min_w), max_w)


def generate_market_report(report: MarketReport, budget: list[tuple],
                           output_path: str = "") -> str:
    """Generate a 5-tab Market Finder Excel workbook."""
    wb = Workbook()

    # ── Tab 1: Executive Summary ──────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"
    ws.cell(row=1, column=1, value="Market Analysis Report").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"County: {report.county}").font = _SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Date: {report.analysis_date}").font = _LABEL_FONT

    ws.cell(row=5, column=1, value="Top 10 Target Zip Codes").font = _SUBTITLE_FONT
    _write_headers(ws, 6, ["Rank", "ZIP", "County", "Grade", "Score",
                            "Total Notices", "Median Value", "Avg Equity %"])
    for i, z in enumerate(report.top_zips[:10], 7):
        ws.cell(row=i, column=1, value=z.rank)
        ws.cell(row=i, column=2, value=z.zip_code)
        ws.cell(row=i, column=3, value=z.county)
        grade_cell = ws.cell(row=i, column=4, value=z.grade)
        grade_cell.fill = _GRADE_COLORS.get(z.grade, PatternFill())
        ws.cell(row=i, column=5, value=round(z.score, 1))
        ws.cell(row=i, column=6, value=z.total_notices)
        ws.cell(row=i, column=7, value=round(z.median_value)).number_format = _MONEY_FMT
        ws.cell(row=i, column=8, value=f"{z.avg_equity_pct:.0f}%" if z.avg_equity_pct else "")
        for c in range(1, 9):
            ws.cell(row=i, column=c).border = _THIN_BORDER

    summary_row = 18
    summary_data = [
        ("Total Zip Codes Analyzed", str(report.total_zips)),
        ("Total Distress Notices", str(report.total_notices)),
        ("Avg Median Home Value", f"${report.avg_median_value:,.0f}" if report.avg_median_value else "N/A"),
        ("Avg Equity", f"{report.avg_equity:.0f}%" if report.avg_equity else "N/A"),
    ]
    for label, value in summary_data:
        ws.cell(row=summary_row, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=summary_row, column=2, value=value).font = _VALUE_FONT
        summary_row += 1

    _auto_widths(ws)

    # ── Tab 2: Zip Scorecard ──────────────────────────────────────
    ws2 = wb.create_sheet("Zip Scorecard")
    ws2.cell(row=1, column=1, value="All Zip Codes — Scored & Ranked").font = _TITLE_FONT
    headers = ["Rank", "ZIP", "County", "Grade", "Score", "Total Notices",
               "Foreclosures", "Tax Sales", "Tax Delinquent", "Probate",
               "Evictions", "Code Violations", "Median Value", "Avg Equity %",
               "Avg Tax Delinquent $", "Properties Analyzed"]
    _write_headers(ws2, 3, headers)
    for i, z in enumerate(report.zip_profiles, 4):
        vals = [z.rank, z.zip_code, z.county, z.grade, round(z.score, 1),
                z.total_notices, z.foreclosure_count, z.tax_sale_count,
                z.tax_delinquent_count, z.probate_count, z.eviction_count,
                z.code_violation_count, round(z.median_value),
                f"{z.avg_equity_pct:.0f}%" if z.avg_equity_pct else "",
                round(z.avg_tax_delinquent_amount), z.property_count]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            if col == 4:
                cell.fill = _GRADE_COLORS.get(val, PatternFill())
            if col == 13:
                cell.number_format = _MONEY_FMT
            cell.border = _THIN_BORDER
    _auto_widths(ws2)

    # ── Tab 3: Distress Density ───────────────────────────────────
    ws3 = wb.create_sheet("Distress Density")
    ws3.cell(row=1, column=1, value="Notice Distribution by Type & ZIP").font = _TITLE_FONT
    _write_headers(ws3, 3, ["ZIP", "County", "Foreclosure", "Tax Sale", "Tax Delinquent",
                             "Probate", "Eviction", "Code Violation", "TOTAL"])
    for i, z in enumerate(sorted(report.zip_profiles, key=lambda x: x.total_notices, reverse=True), 4):
        vals = [z.zip_code, z.county, z.foreclosure_count, z.tax_sale_count,
                z.tax_delinquent_count, z.probate_count, z.eviction_count,
                z.code_violation_count, z.total_notices]
        for col, val in enumerate(vals, 1):
            ws3.cell(row=i, column=col, value=val).border = _THIN_BORDER
    _auto_widths(ws3)

    # ── Tab 4: Competition Map ────────────────────────────────────
    ws4 = wb.create_sheet("Competition Map")
    ws4.cell(row=1, column=1, value="Investor Activity by ZIP").font = _TITLE_FONT
    ws4.cell(row=2, column=1,
             value="Lower competition ratio = less investor activity = better opportunity").font = _LABEL_FONT
    _write_headers(ws4, 4, ["ZIP", "County", "Investor Purchases", "Competition Ratio", "Grade"])
    for i, z in enumerate(sorted(report.zip_profiles, key=lambda x: x.competition_ratio), 5):
        ws4.cell(row=i, column=1, value=z.zip_code)
        ws4.cell(row=i, column=2, value=z.county)
        ws4.cell(row=i, column=3, value=z.investor_purchase_count)
        ws4.cell(row=i, column=4, value=f"{z.competition_ratio:.1%}" if z.competition_ratio else "N/A")
        ws4.cell(row=i, column=5, value=z.grade).fill = _GRADE_COLORS.get(z.grade, PatternFill())
        for c in range(1, 6):
            ws4.cell(row=i, column=c).border = _THIN_BORDER
    _auto_widths(ws4)

    # ── Tab 5: Budget Recommendations ─────────────────────────────
    ws5 = wb.create_sheet("Recommendations")
    ws5.cell(row=1, column=1, value="Marketing Budget Allocation").font = _TITLE_FONT
    ws5.cell(row=2, column=1, value="Budget weighted by zip score — higher-scoring zips get more spend").font = _LABEL_FONT

    _write_headers(ws5, 4, ["ZIP", "Monthly Budget", "Grade", "% of Total"])
    total_budget = sum(b[1] for b in budget) if budget else 1
    for i, (zip_code, amount, grade) in enumerate(budget, 5):
        ws5.cell(row=i, column=1, value=zip_code)
        ws5.cell(row=i, column=2, value=amount).number_format = _MONEY_FMT
        ws5.cell(row=i, column=3, value=grade).fill = _GRADE_COLORS.get(grade, PatternFill())
        ws5.cell(row=i, column=4, value=f"{amount / total_budget:.0%}")
        for c in range(1, 5):
            ws5.cell(row=i, column=c).border = _THIN_BORDER
    _auto_widths(ws5)

    # Save
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(config.OUTPUT_DIR / f"market_analysis_{report.county}_{timestamp}.xlsx")

    wb.save(output_path)
    logger.info("Market report saved to %s", output_path)
    return output_path


# ── Main entry point ──────────────────────────────────────────────────

def run_market_analysis(counties: list[str] | None = None,
                        zip_codes: list[str] | None = None,
                        monthly_budget: float = 5000.0,
                        output_path: str = "") -> dict:
    """Run market analysis for specified counties.

    Returns dict with report data and output path.
    """
    counties = counties or ["Knox", "Blount"]
    county_str = ", ".join(counties)
    logger.info("Starting market analysis for: %s", county_str)

    # Step 1: Load and aggregate our own notice data
    profiles = _load_notice_data(counties)

    # If specific zips requested, filter
    if zip_codes:
        profiles = {k: v for k, v in profiles.items() if k in zip_codes}

    # Add known zips that may not have notices yet
    if not zip_codes:
        for county in counties:
            for z in COUNTY_ZIPS.get(county.lower(), []):
                if z not in profiles:
                    profiles[z] = ZipProfile(zip_code=z, county=county.title())

    if not profiles:
        logger.warning("No data found for analysis")
        return {"error": "No data found"}

    logger.info("Loaded notice data for %d zip codes", len(profiles))

    # Step 2: Score and rank
    scored = score_zip_codes(profiles)
    top = scored[:10]

    # Step 3: Budget allocation
    budget = _allocate_budget(scored, monthly_budget)

    # Step 4: Build report
    values = [z.median_value for z in scored if z.median_value > 0]
    equities = [z.avg_equity_pct for z in scored if z.avg_equity_pct > 0]

    report = MarketReport(
        county=county_str,
        analysis_date=datetime.now().strftime("%Y-%m-%d"),
        total_zips=len(scored),
        zip_profiles=scored,
        top_zips=top,
        total_notices=sum(z.total_notices for z in scored),
        avg_median_value=sum(values) / len(values) if values else 0,
        avg_equity=sum(equities) / len(equities) if equities else 0,
        total_distress=sum(z.total_notices for z in scored),
    )

    # Step 5: Generate Excel report
    report_path = generate_market_report(report, budget, output_path)

    logger.info("Market analysis complete: %d zips scored, top zip %s (score %.1f, grade %s)",
                len(scored), top[0].zip_code if top else "N/A",
                top[0].score if top else 0, top[0].grade if top else "N/A")

    return {
        "report": report,
        "budget": budget,
        "report_path": report_path,
    }
