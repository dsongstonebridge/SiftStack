"""Room-by-room rehab cost estimation with 4-tier finish system.

Generates full rehab budgets, wholetail comparisons, and project timelines.
Regional pricing calibrated for Knoxville / East Tennessee market.

Usage:
  python src/main.py rehab --address "123 Main St, Knoxville, TN 37918"
  python src/main.py rehab --address "123 Main St" --tier 2 --scope full --region knoxville
"""

import logging
from dataclasses import dataclass, field
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config

logger = logging.getLogger(__name__)

# ── Regional multipliers ──────────────────────────────────────────────
# Regional labor/materials costs relative to national average (1.0)
REGIONAL_MULTIPLIERS = {
    "knoxville": 0.88,
    "blount": 0.86,
    "national": 1.00,
    "nashville": 0.95,
    "chattanooga": 0.90,
    "tulsa": 0.82,
    "broken arrow": 0.82,
}
DEFAULT_REGION = ""

# ── 4-Tier Finish System ─────────────────────────────────────────────
# Cost per sqft by tier (national average, before regional multiplier)
TIER_NAMES = {
    1: "Minimum Viable",     # Cheapest materials, basic function
    2: "Builder Grade",      # Standard new construction level
    3: "Mid-Range Upgrade",  # Granite, hardwood, updated fixtures
    4: "Premium/Custom",     # High-end finishes, custom work
}

# ── Room cost tables ──────────────────────────────────────────────────
# Each room category has cost ranges per tier: {tier: (materials, labor)}
# All costs in USD, national average, before regional multiplier

KITCHEN_COSTS = {
    1: {"demo": 500, "cabinets": 2500, "countertops": 800, "appliances": 1500,
        "fixtures": 300, "backsplash": 0, "flooring": 0, "paint": 200, "labor": 2000},
    2: {"demo": 750, "cabinets": 5000, "countertops": 2000, "appliances": 3000,
        "fixtures": 500, "backsplash": 500, "flooring": 0, "paint": 300, "labor": 4000},
    3: {"demo": 1000, "cabinets": 8000, "countertops": 4000, "appliances": 5000,
        "fixtures": 800, "backsplash": 1200, "flooring": 0, "paint": 400, "labor": 6000},
    4: {"demo": 1500, "cabinets": 15000, "countertops": 8000, "appliances": 8000,
        "fixtures": 1500, "backsplash": 2500, "flooring": 0, "paint": 500, "labor": 10000},
}

MASTER_BATH_COSTS = {
    1: {"demo": 400, "vanity": 300, "toilet": 150, "tub_shower": 400, "tile": 500,
        "fixtures": 200, "paint": 150, "labor": 1500},
    2: {"demo": 600, "vanity": 800, "toilet": 250, "tub_shower": 1200, "tile": 1500,
        "fixtures": 400, "paint": 200, "labor": 3000},
    3: {"demo": 800, "vanity": 1500, "toilet": 400, "tub_shower": 2500, "tile": 3000,
        "fixtures": 800, "paint": 250, "labor": 5000},
    4: {"demo": 1000, "vanity": 3000, "toilet": 600, "tub_shower": 5000, "tile": 5000,
        "fixtures": 1500, "paint": 300, "labor": 8000},
}

SECONDARY_BATH_COSTS = {
    1: {"demo": 300, "vanity": 200, "toilet": 150, "tub_shower": 350, "tile": 400,
        "fixtures": 150, "paint": 100, "labor": 1200},
    2: {"demo": 500, "vanity": 500, "toilet": 250, "tub_shower": 800, "tile": 1000,
        "fixtures": 300, "paint": 150, "labor": 2500},
    3: {"demo": 700, "vanity": 1000, "toilet": 350, "tub_shower": 1800, "tile": 2000,
        "fixtures": 500, "paint": 200, "labor": 3500},
    4: {"demo": 900, "vanity": 2000, "toilet": 500, "tub_shower": 3500, "tile": 3500,
        "fixtures": 1000, "paint": 250, "labor": 6000},
}

# Per-sqft costs for whole-house items
FLOORING_PER_SQFT = {1: 2.50, 2: 4.50, 3: 7.00, 4: 12.00}   # materials
FLOORING_LABOR_PER_SQFT = {1: 1.50, 2: 2.00, 3: 2.50, 4: 3.50}
PAINT_PER_SQFT = {1: 0.50, 2: 0.75, 3: 1.00, 4: 1.50}         # materials (walls + trim)
PAINT_LABOR_PER_SQFT = {1: 0.80, 2: 1.00, 3: 1.20, 4: 1.50}

# Fixed-cost items
WINDOWS_PER_UNIT = {1: 250, 2: 400, 3: 650, 4: 1000}  # per window (materials + labor)
ROOF_PER_SQFT = {1: 3.50, 2: 5.00, 3: 7.00, 4: 10.00}  # per roof sqft (≈ floor sqft × 1.1)
HVAC_COSTS = {1: 4000, 2: 6000, 3: 8500, 4: 12000}     # full system replacement
ELECTRICAL_COSTS = {1: 2000, 2: 4000, 3: 7000, 4: 12000}  # panel + rewire
PLUMBING_COSTS = {1: 1500, 2: 3000, 3: 5000, 4: 8000}    # repipe + fixtures
FOUNDATION_COSTS = {1: 2000, 2: 5000, 3: 10000, 4: 20000}  # structural repair

EXTERIOR_COSTS = {
    1: {"siding": 0, "paint": 1500, "landscaping": 500, "driveway": 0, "labor": 1000},
    2: {"siding": 3000, "paint": 2500, "landscaping": 1500, "driveway": 0, "labor": 2000},
    3: {"siding": 6000, "paint": 4000, "landscaping": 3000, "driveway": 2000, "labor": 4000},
    4: {"siding": 12000, "paint": 6000, "landscaping": 6000, "driveway": 5000, "labor": 7000},
}

# Timeline estimates (weeks) per category per tier
TIMELINE_WEEKS = {
    "kitchen": {1: 1, 2: 2, 3: 3, 4: 5},
    "bathrooms": {1: 1, 2: 2, 3: 3, 4: 4},
    "flooring": {1: 1, 2: 1, 3: 2, 4: 2},
    "paint": {1: 1, 2: 1, 3: 1, 4: 2},
    "windows": {1: 0.5, 2: 1, 3: 1, 4: 2},
    "roof": {1: 1, 2: 1, 3: 1.5, 4: 2},
    "hvac": {1: 0.5, 2: 1, 3: 1, 4: 1.5},
    "electrical": {1: 0.5, 2: 1, 3: 1.5, 4: 2},
    "plumbing": {1: 0.5, 2: 1, 3: 1.5, 4: 2},
    "foundation": {1: 1, 2: 2, 3: 3, 4: 4},
    "exterior": {1: 1, 2: 2, 3: 3, 4: 4},
}

# ── Data structures ───────────────────────────────────────────────────


@dataclass
class RoomEstimate:
    """Cost estimate for a single room/category."""
    category: str = ""
    tier: int = 2
    materials: float = 0.0
    labor: float = 0.0
    total: float = 0.0
    line_items: dict = field(default_factory=dict)
    weeks: float = 0.0
    notes: str = ""


@dataclass
class RehabEstimate:
    """Full rehab estimate for a property."""
    address: str = ""
    tier: int = 2
    scope: str = "full"  # "full" or "wholetail"
    region: str = DEFAULT_REGION
    regional_multiplier: float = 0.88
    sqft: int = 0
    bedrooms: int = 0
    bathrooms: float = 0.0
    year_built: int = 0
    rooms: list = field(default_factory=list)
    total_materials: float = 0.0
    total_labor: float = 0.0
    total_cost: float = 0.0
    total_weeks: float = 0.0
    permits_cost: float = 0.0
    contingency_pct: float = 0.10  # 10% contingency
    contingency_cost: float = 0.0
    grand_total: float = 0.0


# ── Estimation engine ─────────────────────────────────────────────────

def _calc_room(category: str, cost_table: dict, tier: int, multiplier: float,
               quantity: int = 1) -> RoomEstimate:
    """Calculate cost for a room category from its cost table."""
    tier_costs = cost_table.get(tier, cost_table.get(2, {}))
    labor = tier_costs.get("labor", 0) * multiplier * quantity
    materials = sum(v for k, v in tier_costs.items() if k != "labor") * multiplier * quantity

    return RoomEstimate(
        category=category,
        tier=tier,
        materials=round(materials),
        labor=round(labor),
        total=round(materials + labor),
        line_items={k: round(v * multiplier * quantity) for k, v in tier_costs.items()},
        weeks=TIMELINE_WEEKS.get(category.lower().split()[0], {}).get(tier, 1) * quantity,
    )


def _calc_per_sqft(category: str, sqft: int, mat_table: dict, labor_table: dict,
                   tier: int, multiplier: float, timeline_key: str = "") -> RoomEstimate:
    """Calculate cost for a per-sqft category."""
    mat_rate = mat_table.get(tier, mat_table.get(2, 0))
    labor_rate = labor_table.get(tier, labor_table.get(2, 0))
    materials = round(sqft * mat_rate * multiplier)
    labor = round(sqft * labor_rate * multiplier)

    return RoomEstimate(
        category=category,
        tier=tier,
        materials=materials,
        labor=labor,
        total=materials + labor,
        line_items={"materials_per_sqft": round(mat_rate * multiplier, 2),
                    "labor_per_sqft": round(labor_rate * multiplier, 2),
                    "sqft": sqft},
        weeks=TIMELINE_WEEKS.get(timeline_key or category.lower(), {}).get(tier, 1),
    )


def _calc_fixed(category: str, cost_table: dict, tier: int, multiplier: float,
                timeline_key: str = "") -> RoomEstimate:
    """Calculate cost for a fixed-cost category (HVAC, electrical, etc.)."""
    total = round(cost_table.get(tier, cost_table.get(2, 0)) * multiplier)
    # Rough 60/40 labor/materials split for mechanical work
    labor = round(total * 0.6)
    materials = total - labor

    return RoomEstimate(
        category=category,
        tier=tier,
        materials=materials,
        labor=labor,
        total=total,
        line_items={"total_installed": total},
        weeks=TIMELINE_WEEKS.get(timeline_key or category.lower(), {}).get(tier, 1),
    )


def estimate_rehab(address: str = "", sqft: int = 0, bedrooms: int = 3,
                   bathrooms: float = 2.0, year_built: int = 0,
                   tier: int = 2, scope: str = "full",
                   region: str = DEFAULT_REGION) -> RehabEstimate:
    """Generate a full rehab estimate for a property.

    Args:
        address: Property address
        sqft: Living square footage
        bedrooms: Number of bedrooms
        bathrooms: Number of bathrooms (e.g. 2.5)
        year_built: Year built (affects which systems need replacement)
        tier: Finish tier 1-4
        scope: "full" (everything) or "wholetail" (cosmetic only)
        region: Regional pricing key
    """
    multiplier = REGIONAL_MULTIPLIERS.get(region.lower(), REGIONAL_MULTIPLIERS["national"])
    tier = max(1, min(4, tier))

    # Default sqft if not provided
    if not sqft:
        sqft = 1500  # Knoxville average for older SFH

    full_baths = int(bathrooms)
    secondary_baths = max(0, full_baths - 1)
    # Window estimate: ~1 per 100 sqft
    window_count = max(8, sqft // 100)

    rooms = []

    # ── Always included (both wholetail and full) ─────────────────
    # Kitchen
    rooms.append(_calc_room("Kitchen", KITCHEN_COSTS, tier, multiplier))

    # Master bath
    rooms.append(_calc_room("Master Bathroom", MASTER_BATH_COSTS, tier, multiplier))

    # Secondary baths
    if secondary_baths > 0:
        rooms.append(_calc_room("Secondary Bathroom(s)", SECONDARY_BATH_COSTS,
                                tier, multiplier, quantity=secondary_baths))

    # Flooring (whole house)
    rooms.append(_calc_per_sqft("Flooring", sqft, FLOORING_PER_SQFT,
                                FLOORING_LABOR_PER_SQFT, tier, multiplier, "flooring"))

    # Paint (whole house)
    rooms.append(_calc_per_sqft("Paint (Interior)", sqft, PAINT_PER_SQFT,
                                PAINT_LABOR_PER_SQFT, tier, multiplier, "paint"))

    # Exterior
    rooms.append(_calc_room("Exterior", EXTERIOR_COSTS, tier, multiplier))

    # ── Full rehab only ───────────────────────────────────────────
    if scope == "full":
        # Windows
        window_total = round(WINDOWS_PER_UNIT.get(tier, 400) * window_count * multiplier)
        rooms.append(RoomEstimate(
            category="Windows", tier=tier,
            materials=round(window_total * 0.6), labor=round(window_total * 0.4),
            total=window_total,
            line_items={"per_window": round(WINDOWS_PER_UNIT.get(tier, 400) * multiplier),
                        "count": window_count},
            weeks=TIMELINE_WEEKS["windows"].get(tier, 1),
        ))

        # Roof (sqft × 1.1 for slope)
        roof_sqft = int(sqft * 1.1)
        roof_total = round(ROOF_PER_SQFT.get(tier, 5) * roof_sqft * multiplier)
        rooms.append(RoomEstimate(
            category="Roof", tier=tier,
            materials=round(roof_total * 0.5), labor=round(roof_total * 0.5),
            total=roof_total,
            line_items={"per_sqft": round(ROOF_PER_SQFT.get(tier, 5) * multiplier, 2),
                        "roof_sqft": roof_sqft},
            weeks=TIMELINE_WEEKS["roof"].get(tier, 1),
        ))

        # HVAC
        rooms.append(_calc_fixed("HVAC", HVAC_COSTS, tier, multiplier, "hvac"))

        # Electrical
        rooms.append(_calc_fixed("Electrical", ELECTRICAL_COSTS, tier, multiplier, "electrical"))

        # Plumbing
        rooms.append(_calc_fixed("Plumbing", PLUMBING_COSTS, tier, multiplier, "plumbing"))

        # Foundation/Structural (only for older homes)
        if year_built and year_built < 1970:
            rooms.append(_calc_fixed("Foundation/Structural", FOUNDATION_COSTS,
                                     tier, multiplier, "foundation"))

    # ── Totals ────────────────────────────────────────────────────
    total_materials = sum(r.materials for r in rooms)
    total_labor = sum(r.labor for r in rooms)
    total_cost = total_materials + total_labor
    # Parallel work reduces timeline: ~60% of sequential sum
    total_weeks = sum(r.weeks for r in rooms) * 0.6

    permits = round(total_cost * 0.03)  # ~3% for permits
    contingency = round(total_cost * 0.10)  # 10% contingency
    grand_total = total_cost + permits + contingency

    estimate = RehabEstimate(
        address=address,
        tier=tier,
        scope=scope,
        region=region,
        regional_multiplier=multiplier,
        sqft=sqft,
        bedrooms=bedrooms,
        bathrooms=bathrooms,
        year_built=year_built,
        rooms=rooms,
        total_materials=round(total_materials),
        total_labor=round(total_labor),
        total_cost=round(total_cost),
        total_weeks=round(total_weeks, 1),
        permits_cost=permits,
        contingency_cost=contingency,
        grand_total=round(grand_total),
    )

    logger.info("Rehab estimate for %s: %s scope, Tier %d (%s), Total $%s, ~%.0f weeks",
                address or "property", scope, tier, TIER_NAMES[tier],
                f"{grand_total:,.0f}", total_weeks)

    return estimate


def estimate_wholetail(address: str = "", sqft: int = 0, bedrooms: int = 3,
                       bathrooms: float = 2.0, year_built: int = 0,
                       tier: int = 2, region: str = DEFAULT_REGION) -> RehabEstimate:
    """Generate a wholetail (cosmetic-only) estimate."""
    return estimate_rehab(address, sqft, bedrooms, bathrooms, year_built,
                          tier=min(tier, 2), scope="wholetail", region=region)


# ── Excel report generation ──────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_VALUE_FONT = Font(name="Calibri", bold=True, size=13, color="222222")
_MONEY_FMT = '#,##0'
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _fmt(val: float) -> str:
    return f"${val:,.0f}"


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_col_widths(ws, min_w=12, max_w=35):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)


def generate_rehab_report(full_est: RehabEstimate, wholetail_est: RehabEstimate | None = None,
                          output_path: str = "") -> str:
    """Generate a 9-tab Excel workbook rehab estimate report."""
    wb = Workbook()

    # ── Tab 1: Executive Summary ──────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"
    ws.cell(row=1, column=1, value="Rehab Cost Estimate").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=full_est.address or "Subject Property").font = _SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = _LABEL_FONT

    row = 5
    data = [
        ("Finish Tier", f"Tier {full_est.tier} — {TIER_NAMES[full_est.tier]}"),
        ("Region", f"{full_est.region.title()} (×{full_est.regional_multiplier:.2f})"),
        ("Property Size", f"{full_est.sqft:,} sqft"),
        ("Bed/Bath", f"{full_est.bedrooms}bd / {full_est.bathrooms}ba"),
        ("Year Built", str(full_est.year_built) if full_est.year_built else "N/A"),
        ("", ""),
        ("FULL REHAB", ""),
        ("Materials", _fmt(full_est.total_materials)),
        ("Labor", _fmt(full_est.total_labor)),
        ("Subtotal", _fmt(full_est.total_cost)),
        ("Permits (~3%)", _fmt(full_est.permits_cost)),
        ("Contingency (10%)", _fmt(full_est.contingency_cost)),
        ("GRAND TOTAL", _fmt(full_est.grand_total)),
        ("Est. Timeline", f"{full_est.total_weeks:.0f} weeks"),
    ]

    if wholetail_est:
        data += [
            ("", ""),
            ("WHOLETAIL", ""),
            ("Materials", _fmt(wholetail_est.total_materials)),
            ("Labor", _fmt(wholetail_est.total_labor)),
            ("GRAND TOTAL", _fmt(wholetail_est.grand_total)),
            ("Est. Timeline", f"{wholetail_est.total_weeks:.0f} weeks"),
            ("", ""),
            ("SAVINGS (Wholetail vs Full)", _fmt(full_est.grand_total - wholetail_est.grand_total)),
        ]

    for label, value in data:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = _VALUE_FONT
        if label in ("GRAND TOTAL", "FULL REHAB", "WHOLETAIL"):
            cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35

    # ── Tab 2: Wholetail Scope ────────────────────────────────────
    ws2 = wb.create_sheet("Wholetail Scope")
    ws2.cell(row=1, column=1, value="Wholetail Scope (Cosmetic Only)").font = _TITLE_FONT
    ws2.cell(row=2, column=1, value="Paint, flooring, fixtures, landscaping, deep clean").font = _LABEL_FONT

    est = wholetail_est or estimate_wholetail(
        full_est.address, full_est.sqft, full_est.bedrooms,
        full_est.bathrooms, full_est.year_built, region=full_est.region)

    headers = ["Category", "Materials", "Labor", "Total", "Weeks"]
    _write_header_row(ws2, 4, headers)
    for i, room in enumerate(est.rooms, 5):
        ws2.cell(row=i, column=1, value=room.category)
        ws2.cell(row=i, column=2, value=room.materials).number_format = _MONEY_FMT
        ws2.cell(row=i, column=3, value=room.labor).number_format = _MONEY_FMT
        ws2.cell(row=i, column=4, value=room.total).number_format = _MONEY_FMT
        ws2.cell(row=i, column=5, value=room.weeks)
        for c in range(1, 6):
            ws2.cell(row=i, column=c).border = _THIN_BORDER
    total_row = 5 + len(est.rooms)
    ws2.cell(row=total_row, column=1, value="TOTAL").font = _VALUE_FONT
    ws2.cell(row=total_row, column=4, value=est.grand_total).number_format = _MONEY_FMT
    ws2.cell(row=total_row, column=4).font = _VALUE_FONT
    _auto_col_widths(ws2)

    # ── Tab 3: Full Rehab Scope ───────────────────────────────────
    ws3 = wb.create_sheet("Full Rehab Scope")
    ws3.cell(row=1, column=1, value="Full Rehab Scope").font = _TITLE_FONT

    _write_header_row(ws3, 3, headers)
    for i, room in enumerate(full_est.rooms, 4):
        ws3.cell(row=i, column=1, value=room.category)
        ws3.cell(row=i, column=2, value=room.materials).number_format = _MONEY_FMT
        ws3.cell(row=i, column=3, value=room.labor).number_format = _MONEY_FMT
        ws3.cell(row=i, column=4, value=room.total).number_format = _MONEY_FMT
        ws3.cell(row=i, column=5, value=room.weeks)
        for c in range(1, 6):
            ws3.cell(row=i, column=c).border = _THIN_BORDER
    total_row = 4 + len(full_est.rooms)
    ws3.cell(row=total_row, column=1, value="Subtotal").font = _VALUE_FONT
    ws3.cell(row=total_row, column=4, value=full_est.total_cost).number_format = _MONEY_FMT
    ws3.cell(row=total_row + 1, column=1, value="Permits (3%)")
    ws3.cell(row=total_row + 1, column=4, value=full_est.permits_cost).number_format = _MONEY_FMT
    ws3.cell(row=total_row + 2, column=1, value="Contingency (10%)")
    ws3.cell(row=total_row + 2, column=4, value=full_est.contingency_cost).number_format = _MONEY_FMT
    ws3.cell(row=total_row + 3, column=1, value="GRAND TOTAL").font = _VALUE_FONT
    ws3.cell(row=total_row + 3, column=4, value=full_est.grand_total).number_format = _MONEY_FMT
    ws3.cell(row=total_row + 3, column=4).font = _VALUE_FONT
    _auto_col_widths(ws3)

    # ── Tab 4: Room-by-Room Detail ────────────────────────────────
    ws4 = wb.create_sheet("Room-by-Room Detail")
    ws4.cell(row=1, column=1, value="Room-by-Room Line Items").font = _TITLE_FONT

    row = 3
    for room in full_est.rooms:
        ws4.cell(row=row, column=1, value=room.category).font = _SUBTITLE_FONT
        row += 1
        for item, cost in room.line_items.items():
            ws4.cell(row=row, column=1, value=f"  {item.replace('_', ' ').title()}")
            ws4.cell(row=row, column=2, value=cost).number_format = _MONEY_FMT
            ws4.cell(row=row, column=2).border = _THIN_BORDER
            row += 1
        ws4.cell(row=row, column=1, value=f"  Room Total").font = _VALUE_FONT
        ws4.cell(row=row, column=2, value=room.total).number_format = _MONEY_FMT
        ws4.cell(row=row, column=2).font = _VALUE_FONT
        row += 2

    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 15

    # ── Tab 5: Materials List ─────────────────────────────────────
    ws5 = wb.create_sheet("Materials List")
    ws5.cell(row=1, column=1, value="Materials Summary").font = _TITLE_FONT
    _write_header_row(ws5, 3, ["Category", "Materials Cost"])
    for i, room in enumerate(full_est.rooms, 4):
        ws5.cell(row=i, column=1, value=room.category)
        ws5.cell(row=i, column=2, value=room.materials).number_format = _MONEY_FMT
    total_row = 4 + len(full_est.rooms)
    ws5.cell(row=total_row, column=1, value="TOTAL MATERIALS").font = _VALUE_FONT
    ws5.cell(row=total_row, column=2, value=full_est.total_materials).number_format = _MONEY_FMT
    _auto_col_widths(ws5)

    # ── Tab 6: Labor Breakdown ────────────────────────────────────
    ws6 = wb.create_sheet("Labor Breakdown")
    ws6.cell(row=1, column=1, value="Labor Summary").font = _TITLE_FONT
    _write_header_row(ws6, 3, ["Category", "Labor Cost", "Est. Weeks"])
    for i, room in enumerate(full_est.rooms, 4):
        ws6.cell(row=i, column=1, value=room.category)
        ws6.cell(row=i, column=2, value=room.labor).number_format = _MONEY_FMT
        ws6.cell(row=i, column=3, value=room.weeks)
    total_row = 4 + len(full_est.rooms)
    ws6.cell(row=total_row, column=1, value="TOTAL LABOR").font = _VALUE_FONT
    ws6.cell(row=total_row, column=2, value=full_est.total_labor).number_format = _MONEY_FMT
    _auto_col_widths(ws6)

    # ── Tab 7: Timeline ───────────────────────────────────────────
    ws7 = wb.create_sheet("Timeline")
    ws7.cell(row=1, column=1, value="Project Timeline").font = _TITLE_FONT
    ws7.cell(row=2, column=1,
             value=f"Estimated {full_est.total_weeks:.0f} weeks (with parallel work)").font = _LABEL_FONT

    _write_header_row(ws7, 4, ["Category", "Duration (weeks)", "Phase"])
    phase_map = {
        "Foundation/Structural": "1 — Structural",
        "Roof": "1 — Structural",
        "Plumbing": "2 — Mechanical",
        "Electrical": "2 — Mechanical",
        "HVAC": "2 — Mechanical",
        "Windows": "2 — Mechanical",
        "Kitchen": "3 — Interior",
        "Master Bathroom": "3 — Interior",
        "Secondary Bathroom(s)": "3 — Interior",
        "Flooring": "4 — Finish",
        "Paint (Interior)": "4 — Finish",
        "Exterior": "5 — Exterior",
    }
    for i, room in enumerate(full_est.rooms, 5):
        ws7.cell(row=i, column=1, value=room.category)
        ws7.cell(row=i, column=2, value=room.weeks)
        ws7.cell(row=i, column=3, value=phase_map.get(room.category, ""))
    _auto_col_widths(ws7)

    # ── Tab 8: Comparison ─────────────────────────────────────────
    ws8 = wb.create_sheet("Comparison")
    ws8.cell(row=1, column=1, value="Wholetail vs Full Rehab Comparison").font = _TITLE_FONT

    wt = wholetail_est or est
    comp_data = [
        ("", "Wholetail", "Full Rehab", "Difference"),
        ("Materials", wt.total_materials, full_est.total_materials,
         full_est.total_materials - wt.total_materials),
        ("Labor", wt.total_labor, full_est.total_labor,
         full_est.total_labor - wt.total_labor),
        ("Permits", wt.permits_cost, full_est.permits_cost,
         full_est.permits_cost - wt.permits_cost),
        ("Contingency", wt.contingency_cost, full_est.contingency_cost,
         full_est.contingency_cost - wt.contingency_cost),
        ("GRAND TOTAL", wt.grand_total, full_est.grand_total,
         full_est.grand_total - wt.grand_total),
        ("Timeline (weeks)", wt.total_weeks, full_est.total_weeks,
         full_est.total_weeks - wt.total_weeks),
    ]
    _write_header_row(ws8, 3, comp_data[0])
    for i, (label, wt_val, full_val, diff) in enumerate(comp_data[1:], 4):
        ws8.cell(row=i, column=1, value=label)
        ws8.cell(row=i, column=2, value=wt_val).number_format = _MONEY_FMT
        ws8.cell(row=i, column=3, value=full_val).number_format = _MONEY_FMT
        ws8.cell(row=i, column=4, value=diff).number_format = _MONEY_FMT
        for c in range(1, 5):
            ws8.cell(row=i, column=c).border = _THIN_BORDER
    _auto_col_widths(ws8)

    # ── Tab 9: Notes & Assumptions ────────────────────────────────
    ws9 = wb.create_sheet("Notes & Assumptions")
    ws9.cell(row=1, column=1, value="Notes & Assumptions").font = _TITLE_FONT
    notes = [
        f"Tier {full_est.tier}: {TIER_NAMES[full_est.tier]}",
        f"Regional multiplier: {full_est.region.title()} = {full_est.regional_multiplier:.2f}x national avg",
        "",
        "Tier Definitions:",
        "  Tier 1 (Minimum Viable): Cheapest materials, basic function. Rental-ready.",
        "  Tier 2 (Builder Grade): Standard new construction level. Most common for flips.",
        "  Tier 3 (Mid-Range): Granite, hardwood, updated fixtures. Higher ARV neighborhoods.",
        "  Tier 4 (Premium/Custom): High-end finishes, custom work. Luxury market only.",
        "",
        "Assumptions:",
        "  - Permits estimated at 3% of total cost",
        "  - 10% contingency included for unforeseen issues",
        "  - Timeline assumes parallel work (60% of sequential estimate)",
        "  - Foundation/structural work included only for homes built before 1970",
        "  - Window count estimated at 1 per 100 sqft of living area",
        "  - Roof area estimated at 1.1x living area sqft",
        "",
        "Wholetail Scope includes: Kitchen, Bathrooms, Flooring, Paint, Exterior (cosmetic)",
        "Wholetail Excludes: Windows, Roof, HVAC, Electrical, Plumbing, Foundation",
        "",
        "After calibrating with 3-5 closed deals, estimates typically tighten to",
        "within 10-15% of actual contractor SOWs.",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    for i, note in enumerate(notes, 3):
        ws9.cell(row=i, column=1, value=note).font = _LABEL_FONT
    ws9.column_dimensions["A"].width = 70

    # Save
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_addr = "".join(c if c.isalnum() or c in " -" else "_" for c in (full_est.address or "property"))[:40]
        output_path = str(config.OUTPUT_DIR / f"rehab_estimate_{safe_addr}_{timestamp}.xlsx")

    wb.save(output_path)
    logger.info("Rehab report saved to %s", output_path)
    return output_path


# ── Main entry point ──────────────────────────────────────────────────

def run_rehab_estimate(address: str = "", sqft: int = 0, bedrooms: int = 3,
                       bathrooms: float = 2.0, year_built: int = 0,
                       tier: int = 2, scope: str = "full",
                       region: str = DEFAULT_REGION,
                       output_path: str = "") -> dict:
    """Run rehab estimation and generate report.

    Returns dict with estimates and report path.
    """
    logger.info("Estimating rehab for: %s (%s sqft, Tier %d, %s scope)",
                address or "property", sqft, tier, scope)

    # Full rehab estimate
    full_est = estimate_rehab(address, sqft, bedrooms, bathrooms, year_built,
                              tier=tier, scope="full", region=region)

    # Wholetail comparison
    wholetail_est = estimate_wholetail(address, sqft, bedrooms, bathrooms,
                                       year_built, region=region)

    # Generate report
    report_path = generate_rehab_report(full_est, wholetail_est, output_path)

    return {
        "full_estimate": full_est,
        "wholetail_estimate": wholetail_est,
        "report_path": report_path,
    }
